"""
parse_excel.py - Reads Plan_de_ensayos_2026.xlsx and returns structured data.
Values in cells: '*' = planned, 0 = not done, 0.5 = done but not in repo, 1 = done correctly.
"""

import openpyxl
import json
from pathlib import Path


PROJECTS = [
    'Syrah', 'Luar', 'Gregal', 'El Turpial', 'El Alamo', 'El Guayacan',
    'Pto tranquilo', 'La Alegria IV', 'La Alegria V', 'La Gratitud I',
    'La Gratitud II', 'La Gratirud IV', 'Murales', 'Vizcaya', 'Tusset',
    'Verdemonte', 'Entrebosques VI'
]

# Column index (0-based) where each month block starts
MONTH_START_COLS = {
    1: 5, 2: 22, 3: 39, 4: 56, 5: 73, 6: 90,
    7: 107, 8: 124, 9: 141, 10: 158, 11: 175, 12: 192
}

STATUS_MAP = {
    '*': 'planned',
    0: 'not_done',
    0.5: 'partial',
    1: 'done',
    None: None
}


def parse_excel(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    data = []

    for row_idx in range(3, ws.max_row + 1):
        row = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        etapa = row[0]
        material = row[1]
        ensayo = row[2]
        ntc = row[3]
        frecuencia = row[4]

        if not etapa or not ensayo:
            continue

        record = {
            'etapa': etapa,
            'material': material,
            'ensayo': ensayo,
            'ntc': ntc,
            'frecuencia': frecuencia,
            'schedule': {}
        }

        for month, start_col in MONTH_START_COLS.items():
            month_projects = []
            for p_idx, proj in enumerate(PROJECTS):
                col_val = row[start_col + p_idx]
                if col_val is not None:
                    status = STATUS_MAP.get(col_val, 'planned')
                    month_projects.append({
                        'project': proj,
                        'value': col_val,
                        'status': status
                    })
            if month_projects:
                record['schedule'][str(month)] = month_projects

        data.append(record)

    return data


def get_summary(data: list[dict]) -> dict:
    """Compute summary statistics from parsed data."""
    total_planned = 0
    done = 0
    partial = 0
    not_done = 0
    planned_only = 0

    proj_month: dict[str, dict[str, list]] = {}
    etapa_counts: dict[str, int] = {}
    monthly_totals: dict[str, int] = {str(m): 0 for m in range(1, 13)}
    monthly_done: dict[str, int] = {str(m): 0 for m in range(1, 13)}

    for rec in data:
        etapa = rec['etapa']
        for month, projs in rec['schedule'].items():
            for p in projs:
                proj = p['project']
                status = p['status']
                total_planned += 1
                monthly_totals[month] = monthly_totals.get(month, 0) + 1

                if status == 'done':
                    done += 1
                    monthly_done[month] = monthly_done.get(month, 0) + 1
                elif status == 'partial':
                    partial += 1
                elif status == 'not_done':
                    not_done += 1
                else:
                    planned_only += 1

                if proj not in proj_month:
                    proj_month[proj] = {}
                if month not in proj_month[proj]:
                    proj_month[proj][month] = []
                proj_month[proj][month].append({
                    'ensayo': rec['ensayo'],
                    'etapa': rec['etapa'],
                    'material': rec['material'],
                    'status': status,
                    'value': p['value']
                })

                etapa_counts[etapa] = etapa_counts.get(etapa, 0) + 1

    return {
        'total': total_planned,
        'done': done,
        'partial': partial,
        'not_done': not_done,
        'planned_only': planned_only,
        'proj_month': proj_month,
        'etapa_counts': etapa_counts,
        'monthly_totals': monthly_totals,
        'monthly_done': monthly_done,
    }


if __name__ == '__main__':
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else 'Plan_de_ensayos_2026.xlsx'
    data = parse_excel(path)
    print(json.dumps(data, ensure_ascii=False, indent=2))
