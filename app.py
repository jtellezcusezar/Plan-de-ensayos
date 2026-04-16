import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from pathlib import Path
from calendar import monthrange
from datetime import datetime, timezone
from zoneinfo import ZoneInfo
from io import BytesIO
import base64
import html
from openpyxl import load_workbook
import json
from uuid import uuid4

st.set_page_config(
    page_title="Calidad - Cusezar", page_icon="🏗️",
    layout="wide", initial_sidebar_state="expanded",
)

# ── CONSTANTES ─────────────────────────────────────────────────────────────────
MESES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
         7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
ESTADO_MAP = {"*":"Planeado", "0":"No Realizado", "0,5":"Incompleto", "1":"Completo"}
META = 90

CONTROL_AREA_OPTIONS = ["Torre", "Zonas comunes", "Diseño", "Curado", "Producto terminado"]

PRODUCTO_TERMINADO_ENSAYOS = [
    ("Estructura", "Acero", "Ferroscan"),
    ("Estructura", "Pilotaje", "Pruebas Pit"),
    ("Obra Gris", "Muretes", "Resistencia a la Compresión (muros internos)"),
    ("Obra Gris", "Muretes", "Resistencia a la Compresión (fachadas)"),
    ("Obra Gris", "Muretes", "Resistencia a la Compresión (Mamposteria Estructural)"),
    ("Obra Gris", "Acero", "Ferroscan"),
    ("Obra Blanca", "Ascensores", "Certificado de ascensores"),
    ("Obra Blanca", "Puertas Eléctricas", "Certificado de puertas electricas"),
    ("Obra Blanca", "Equipo de Suministro", "Manografo"),
    ("Obra Blanca", "Equipo de Suministro", "Certificado de lavado de tanques"),
    ("Obra Blanca", "Red Contra Incendios", "Hidrostatica"),
    ("Obra Blanca", "Red Contra Incendios", "Apertura maual de valvulas"),
    ("Obra Blanca", "Red Contra Incendios", "Pitometrica"),
    ("Obra Blanca", "Seguridad y Control", "Señalización e iluminacion ruta de evacuación"),
    ("Obra Blanca", "Electrica", "Apantallamiento"),
    ("Obra Blanca", "Barandas", "Prueba de carga baranda"),
    ("Obra Blanca", "Impermeabilización de Cubierta", "Estanqueidad en placas descubiertas"),
    ("Obra Blanca", "Impermeabilización de Fachada", "Tubo Rilem o pipetas de Karsten."),
    ("Obra Blanca", "Impermeabilización de Fachada", "Prueba de Perlado"),
]

PRODUCTO_TERMINADO_CONTROLES = [
    ("Torre", "Estructura", "Control Instalaciones Sanitarias (Estanqueidad)"),
    ("Torre", "Estructura", "Control Instalaciones Sanitarias (Flujo)"),
    ("Torre", "Obra gris y Obra blanca", "Acta de liberacion Pañete"),
    ("Torre", "Obra gris y Obra blanca", "Acta de liberacion Mamposteria"),
    ("Torre", "Obra gris y Obra blanca", "Afinados de pisos"),
    ("Torre", "Obra gris y Obra blanca", "Control Aparatos Sanitarios"),
    ("Torre", "Obra gris y Obra blanca", "Control de pintura"),
    ("Torre", "Producto terminado", "Control instrumentacion pantallas"),
    ("Torre", "Producto terminado", "Actas de liberacion control Niveles placa techo estructura"),
    ("Torre", "Producto terminado", "Control de Asentamiento"),
    ("Torre", "Producto terminado", "Niveles de ruido"),
    ("Torre", "Producto terminado", "Prueba de estanqueidad, Terrazas"),
    ("Torre", "Producto terminado", "Inspección recepción foso de ascensor"),
]

COLORS = {
    "Planeado":     "#7BA7D4",
    "Completo":     "#6BBF9E",
    "Incompleto":   "#E8C17A",
    "No Realizado": "#D98B8B",
}

BASE_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font=dict(family="Inter, sans-serif", color="#272829", size=12),
    margin=dict(t=40, b=10, l=10, r=10),
    hoverlabel=dict(
        bgcolor="#1E293B",
        font_color="#F1F5F9",
        font_size=12,
        bordercolor="#334155",
    ),
)

def apply_base(fig, h=300, legend_h=True):
    """Aplica el layout base y la leyenda horizontal (por defecto)."""
    fig.update_layout(
        paper_bgcolor=BASE_LAYOUT["paper_bgcolor"],
        plot_bgcolor=BASE_LAYOUT["plot_bgcolor"],
        font=BASE_LAYOUT["font"],
        margin=BASE_LAYOUT["margin"],
        hoverlabel=BASE_LAYOUT["hoverlabel"],
        height=h,
    )
    if legend_h:
        fig.update_layout(
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=1.02,
                font=dict(size=11),
            )
        )
    return fig

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=DM+Mono:wght@400;500&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif!important;}
#MainMenu,footer{visibility:hidden;}
.block-container{padding-top:3.25rem!important;max-width:100%!important;padding-left:2rem!important;padding-right:2rem!important;}
section[data-testid="stSidebar"],[data-testid="stSidebar"]{background:#F8FAFC;border-right:1px solid #E5E9F0;}
section[data-testid="stSidebar"] > div:first-child,[data-testid="stSidebar"] > div:first-child{padding-top:1.25rem;}
section[data-testid="stSidebar"] > div:first-child,[data-testid="stSidebar"] > div:first-child{display:flex;flex-direction:column;height:100%;}
section[data-testid="stSidebar"] [data-testid="stVerticalBlock"],[data-testid="stSidebar"] [data-testid="stVerticalBlock"]{gap:.35rem;}
.sidebar-logo-wrap{display:flex;justify-content:center;align-items:center;width:fit-content;margin:0 auto 14px auto;padding:12px 18px;background:#FF0000;border-radius:18px;}
.sidebar-logo-img{display:block;width:150px;height:auto;}
.sidebar-brand{font-size:18px;font-weight:800;color:#111827;line-height:1.2;margin-bottom:4px;text-align:center;}
.sidebar-sub{font-size:11px;color:#9CA3AF;font-weight:600;text-transform:uppercase;letter-spacing:.05em;margin-bottom:14px;text-align:center;}
.sidebar-footer{margin-top:auto;padding-top:16px;font-size:11px;color:#6B7280;text-align:center;line-height:1.5;}
.sidebar-footer strong{display:block;font-size:10px;color:#9CA3AF;text-transform:uppercase;letter-spacing:.05em;margin-bottom:4px;}
section[data-testid="stSidebar"] .stRadio > label,[data-testid="stSidebar"] .stRadio > label{display:none;}
section[data-testid="stSidebar"] .stRadio [role="radiogroup"],[data-testid="stSidebar"] .stRadio [role="radiogroup"]{gap:8px;}
section[data-testid="stSidebar"] .stRadio [role="radio"],[data-testid="stSidebar"] .stRadio [role="radio"]{background:#FFFFFF;border:1px solid #E5E9F0;border-radius:12px;padding:10px 12px;color:#475569;font-size:13px;font-weight:600;box-shadow:0 1px 3px rgba(15,23,42,.04);}
section[data-testid="stSidebar"] .stRadio [role="radio"][aria-checked="true"],[data-testid="stSidebar"] .stRadio [role="radio"][aria-checked="true"]{background:#EEF3FA;border-color:#C8DCF0;color:#4A7BA8;box-shadow:0 4px 14px rgba(74,123,168,.08);}
section[data-testid="stSidebar"] .stRadio [role="radio"]:hover,[data-testid="stSidebar"] .stRadio [role="radio"]:hover{border-color:#CBD5E1;}
section[data-testid="stSidebar"] [data-testid="stCaptionContainer"],[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p{color:#9CA3AF!important;}
button[kind="header"]{opacity:1!important;visibility:visible!important;}
.page-head{display:flex;align-items:flex-end;justify-content:space-between;gap:16px;margin:0 0 18px 0;padding-bottom:10px;border-bottom:1px solid #E5E9F0;}
.page-title{font-size:24px;font-weight:800;color:#111827;line-height:1.1;margin:0;}
.page-sub{font-size:12px;color:#9CA3AF;margin:4px 0 0 0;}
.filter-bar{margin-bottom:16px;}
.filter-bar-title{display:none;}
div[data-testid="stSelectbox"]>label{font-size:11px!important;font-weight:600!important;color:#6B7280!important;text-transform:uppercase!important;letter-spacing:.05em!important;margin-bottom:4px!important;}
div[data-testid="stSelectbox"]>div>div{border-radius:10px!important;border:1.5px solid #E5E9F0!important;background:#FAFBFC!important;font-size:13px!important;color:#374151!important;}
div[data-testid="stSelectbox"]>div>div:focus-within{border-color:#7BA7D4!important;box-shadow:0 0 0 3px rgba(123,167,212,.12)!important;}
div[data-testid="stTextInput"]>label{font-size:11px!important;font-weight:600!important;color:#6B7280!important;text-transform:uppercase!important;letter-spacing:.05em!important;}
div[data-testid="stTextInput"]>div>input{border-radius:10px!important;border:1.5px solid #E5E9F0!important;background:#FAFBFC!important;font-size:13px!important;color:#374151!important;}
div[data-testid="stTextInput"]>div>input:focus{border-color:#7BA7D4!important;box-shadow:0 0 0 3px rgba(123,167,212,.12)!important;}
.kpi-card{background:#fff;border:1px solid #E5E9F0;border-radius:14px;padding:18px 20px;position:relative;overflow:hidden;box-shadow:0 1px 4px rgba(15,23,42,.05);transition:transform .15s,box-shadow .15s;}
.kpi-card:hover{transform:translateY(-2px);box-shadow:0 6px 20px rgba(15,23,42,.08);}
.kpi-card::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;border-radius:14px 14px 0 0;}
.kp-blue::before{background:#7BA7D4;}.kp-green::before{background:#6BBF9E;}.kp-yellow::before{background:#E8C17A;}.kp-red::before{background:#D98B8B;}.kp-slate::before{background:linear-gradient(90deg,#7BA7D4,#6BBF9E);}
.kpi-icon{font-size:20px;margin-bottom:8px;}.kpi-label{font-size:10px;font-weight:700;color:#9CA3AF;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px;}
.kpi-value{font-size:28px;font-weight:800;line-height:1;font-family:'DM Mono',monospace;margin-bottom:4px;}.kpi-sub{font-size:11px;color:#9CA3AF;}
.kp-blue .kpi-value{color:#4A7BA8;}.kp-green .kpi-value{color:#3D8B6E;}.kp-yellow .kpi-value{color:#C49A3C;}.kp-red .kpi-value{color:#B05B5B;}.kp-slate .kpi-value{color:#4A7BA8;}
.section-head{display:flex;align-items:flex-end;justify-content:space-between;gap:16px;margin:0 0 10px 0;padding-bottom:8px;border-bottom:1px solid #E5E9F0;}
.section-title{font-size:15px;font-weight:700;color:#111827;line-height:1.2;margin:0;}
.section-sub{font-size:11px;color:#9CA3AF;line-height:1.4;text-align:right;max-width:60%;margin:0;}
.dash-card{background:transparent;border:none;border-radius:0;padding:0;margin-bottom:18px;box-shadow:none;}
.card-title{font-size:14px;font-weight:700;color:#111827;margin-bottom:2px;}.card-sub{font-size:11px;color:#9CA3AF;margin-bottom:14px;}
.info-note{padding:9px 14px;background:#EEF3FA;border:1px solid #C8DCF0;border-radius:10px;font-size:12px;color:#4A7BA8;font-weight:500;margin-bottom:18px;}
.ok-note{padding:9px 14px;background:#E4F4EE;border:1px solid #A8D5BF;border-radius:10px;font-size:12px;color:#3D8B6E;font-weight:500;}
.sem-grid{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:20px;}
.sem-card{background:#fff;border:1px solid #E5E9F0;border-radius:12px;padding:14px 18px;min-width:160px;position:relative;overflow:hidden;box-shadow:0 1px 4px rgba(15,23,42,.05);transition:transform .15s;}
.sem-card:hover{transform:translateY(-1px);}
.sem-card::after{content:'';position:absolute;bottom:0;left:0;right:0;height:3px;border-radius:0 0 12px 12px;}
.sv::after{background:#6BBF9E;}.sa::after{background:#E8C17A;}.sr::after{background:#D98B8B;}
.sem-dot{width:8px;height:8px;border-radius:50%;display:inline-block;margin-right:6px;}
.sem-name{font-size:12px;font-weight:700;color:#111827;margin-bottom:6px;display:flex;align-items:center;}
.sem-tasa{font-size:22px;font-weight:800;font-family:'DM Mono',monospace;}
.sv .sem-tasa{color:#3D8B6E;}.sa .sem-tasa{color:#C49A3C;}.sr .sem-tasa{color:#B05B5B;}
.sem-detail{font-size:11px;color:#9CA3AF;margin-top:3px;}
.hm-wrap{overflow-x:auto;border-radius:10px;border:1px solid #E5E9F0;}
.hm-table{width:100%;border-collapse:collapse;font-size:12px;}
.hm-table th{background:#F8F9FB;padding:9px 8px;text-align:center;font-size:10px;font-weight:700;color:#9CA3AF;text-transform:uppercase;border-bottom:1px solid #E5E9F0;white-space:nowrap;}
.hm-table th.hmp{text-align:left;min-width:160px;padding-left:16px;}
.hm-table td{padding:8px;text-align:center;font-weight:700;font-family:'DM Mono',monospace;border-bottom:1px solid #F3F4F6;}
.hm-table td.hmpn{text-align:left;font-family:'Inter',sans-serif;font-size:12px;padding-left:16px;color:#111827;font-weight:600;}
.hm-table tr:last-child td{border-bottom:none;}.hm-table tr:hover td{filter:brightness(.97);}
.h100{background:#B8E4D0;color:#2D6A4F;}.h75{background:#DDE8B2;color:#667A1E;}.h50{background:#F4E1A6;color:#A97B12;}.h25{background:#EEC39F;color:#A45724;}.h0{background:#F0C8C8;color:#8B2B2B;}.hna{background:#F8F9FB;color:#C4CAD4;font-family:'Inter',sans-serif;font-weight:500;font-size:11px;}
.badge{display:inline-flex;align-items:center;gap:4px;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;}
.bc{background:#E4F4EE;color:#3D8B6E;}.bi{background:#FBF3E0;color:#C49A3C;}.bn{background:#F8E8E8;color:#B05B5B;}.bp{background:#EEF3FA;color:#4A7BA8;}
.rt{width:100%;border-collapse:collapse;font-size:13px;}
.rt th{background:#B5545C;padding:10px 14px;text-align:left;font-size:11px;font-weight:700;color:#FFF7F7;text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid #9F434B;white-space:nowrap;}
.rt td{padding:10px 14px;border-bottom:1px solid #F1D9DB;color:#6B7280;}
.rt td:first-child{background:#F8E8E8;color:#8F3942;font-weight:700;}
.rt tr:last-child td{border-bottom:none;}.rt tr:hover td{background:#FCF4F5;}
.rt tr:hover td:first-child{background:#F8E8E8;}
.ig-wrap{display:flex;justify-content:center;}
.ig-table{width:100%;max-width:1180px;border-collapse:collapse;font-size:12px;table-layout:fixed;}
.ig-table th{background:#B5545C;padding:8px 8px;text-align:center;font-size:10px;font-weight:700;color:#FFF7F7;text-transform:uppercase;letter-spacing:.05em;border-bottom:1px solid #9F434B;white-space:normal;line-height:1.2;word-break:break-word;}
.ig-table td{padding:7px 8px;border-bottom:1px solid #F1D9DB;color:#6B7280;text-align:center;line-height:1.15;word-break:break-word;}
.ig-table td:first-child{background:#F8E8E8;color:#8F3942;font-weight:700;text-align:center;}
.ig-table tr.ig-summary-row td{background:#F8E8E8!important;border-top:1px solid #E8CDD0;border-bottom:1px solid #E8CDD0;font-weight:700;color:#8F3942!important;}
.ig-table tr.ig-summary-row td:first-child{background:#F8E8E8!important;color:#8F3942!important;text-transform:uppercase;letter-spacing:.04em;}
.ig-table tr.ig-summary-row td.h100,
.ig-table tr.ig-summary-row td.h75,
.ig-table tr.ig-summary-row td.h50,
.ig-table tr.ig-summary-row td.h25,
.ig-table tr.ig-summary-row td.h0,
.ig-table tr.ig-summary-row td.hna{background:#F8E8E8!important;}
.ig-table tr.ig-corp-row td{background:#E1B7BB!important;border-top:1px solid #C89499;border-bottom:1px solid #C89499;font-weight:800;}
.ig-table tr.ig-corp-row td:first-child{background:#E1B7BB!important;color:#5E1F28!important;text-transform:uppercase;letter-spacing:.05em;}
.ig-table tr.ig-corp-row td.h100,
.ig-table tr.ig-corp-row td.h75,
.ig-table tr.ig-corp-row td.h50,
.ig-table tr.ig-corp-row td.h25,
.ig-table tr.ig-corp-row td.h0,
.ig-table tr.ig-corp-row td.hna{background:#E1B7BB!important;}
.ig-table tr:last-child td{border-bottom:none;}
.hml{display:flex;gap:8px;flex-wrap:wrap;margin-bottom:14px;align-items:center;}
.hml span{font-size:11px;font-weight:600;padding:3px 10px;border-radius:20px;}
div[data-testid="stDownloadButton"] button{background:#EEF3FA!important;color:#4A7BA8!important;border:1.5px solid #C8DCF0!important;border-radius:8px!important;font-size:12px!important;font-weight:600!important;padding:6px 14px!important;}
div[data-testid="stDownloadButton"] button:hover{background:#7BA7D4!important;color:#fff!important;}
::-webkit-scrollbar{width:5px;height:5px;}::-webkit-scrollbar-track{background:transparent;}::-webkit-scrollbar-thumb{background:#E5E9F0;border-radius:3px;}
</style>
""", unsafe_allow_html=True)

# ── DATA ───────────────────────────────────────────────────────────────────────
EXCEL_PATH = Path("Plan_de_ensayos_2026.xlsx")
SIDEBAR_LOGO_PATH = Path("cusezar.png")


def get_excel_signature(path):
    stat = path.stat()
    return stat.st_mtime_ns, stat.st_size


def format_spanish_date(dt_value):
    return f"{dt_value.day} de {MESES[dt_value.month]} del {dt_value.year}"


def get_excel_last_update_text(path):
    wb = load_workbook(path, read_only=True)
    try:
        updated_at = wb.properties.modified or wb.properties.created
    finally:
        wb.close()

    if updated_at is None:
        stat = path.stat()
        updated_at = datetime.fromtimestamp(stat.st_mtime, ZoneInfo("America/Bogota"))
        return format_spanish_date(updated_at)

    if updated_at.tzinfo is None:
        updated_at = updated_at.replace(tzinfo=timezone.utc)

    return format_spanish_date(updated_at.astimezone(ZoneInfo("America/Bogota")))


def read_excel_table(path, table_name):
    """Lee una tabla nombrada de Excel sin depender del nombre de la hoja."""
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        for ws in wb.worksheets:
            if table_name in ws.tables:
                ref = ws.tables[table_name].ref
                rows = list(ws[ref])
                data = [[cell.value for cell in row] for row in rows]
                headers = data[0]
                values = data[1:]
                return pd.DataFrame(values, columns=headers)
    finally:
        wb.close()
    raise ValueError(f"No se encontró la tabla '{table_name}' en {path}.")

def normalize_text_columns(df):
    """Limpia espacios en columnas de texto."""
    df = df.copy()
    for col in df.columns:
        if pd.api.types.is_object_dtype(df[col]):
            df[col] = df[col].astype(str).str.strip()
    return df

def normalize_status_series(series):
    series = series.astype(str).str.strip()
    return series.replace({
        "0.0": "0",
        "1.0": "1",
        "0.5": "0,5",
        "nan": None,
        "None": None,
        "": None,
    })

def parse_text_value(value):
    if value is None or pd.isna(value):
        return None
    txt = str(value).strip()
    if txt in {"", "None", "nan", "*"}:
        return None
    txt = txt.replace("%", "").replace(",", ".")
    try:
        num = float(txt)
    except ValueError:
        return None
    return num * 100 if 0 <= num <= 1 else num


def normalize_project_key(value):
    if value is None or pd.isna(value):
        return None
    text = " ".join(str(value).split()).strip()
    if not text or text.lower() in {"nan", "none"}:
        return None
    return text.casefold()


def normalize_month_header(value):
    if value is None or pd.isna(value):
        return None

    if isinstance(value, datetime):
        return value.month

    if isinstance(value, (int, float)) and not pd.isna(value):
        month_num = int(value)
        if 1 <= month_num <= 12:
            return month_num

    text = str(value).strip()
    if not text:
        return None

    normalized_text = (
        text.casefold()
        .replace(".", "")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )

    month_lookup = {}
    for month, nombre in MESES.items():
        normalized_name = (
            nombre.casefold()
            .replace("á", "a")
            .replace("é", "e")
            .replace("í", "i")
            .replace("ó", "o")
            .replace("ú", "u")
        )
        month_lookup[normalized_name] = month
        month_lookup[normalized_name[:3]] = month

    return month_lookup.get(normalized_text)

@st.cache_data
def load_data(excel_signature):
    df_ensayos = read_excel_table(EXCEL_PATH, "Ensayos")
    df_ensayos = df_ensayos.rename(columns={
        "Etapa": "ETAPA",
        "Material": "MATERIAL",
        "Ensayo": "ENSAYO",
        "Frecuencia": "FRECUENCIA",
        "Valor": "Cantidad",
    })
    df_ensayos = normalize_text_columns(df_ensayos)

    df_controles = read_excel_table(EXCEL_PATH, "Controles")
    df_controles = normalize_text_columns(df_controles)
    if "Valor" in df_controles.columns:
        df_controles["Valor"] = normalize_status_series(df_controles["Valor"])
        df_controles["Valor_num"] = pd.to_numeric(
            df_controles["Valor"].replace({"0,5": "0.5", "*": None}),
            errors="coerce"
        )

    df_ensayos["Cantidad"] = normalize_status_series(df_ensayos["Cantidad"])
    df_ensayos["MesNombre"]    = df_ensayos["Mes"].map(MESES)
    df_ensayos["Estado"]       = df_ensayos["Cantidad"].map(lambda x: ESTADO_MAP.get(x, str(x)))
    df_ensayos["EsEjecutado"]  = df_ensayos["Cantidad"] != "*"
    df_ensayos["Cantidad_num"] = pd.to_numeric(
        df_ensayos["Cantidad"].replace({"0,5": "0.5", "*": None}),
        errors="coerce"
    )

    try:
        df_2025 = read_excel_table(EXCEL_PATH, "Datos_2025")
        df_2025 = normalize_text_columns(df_2025)
    except ValueError:
        df_2025 = pd.DataFrame()

    return df_ensayos, df_controles, df_2025


EXCEL_SIGNATURE = get_excel_signature(EXCEL_PATH)
EXCEL_LAST_UPDATE_TEXT = get_excel_last_update_text(EXCEL_PATH)
df_full, df_controles, df_2025 = load_data(EXCEL_SIGNATURE)
meses_con_datos = sorted(df_full[df_full["EsEjecutado"]]["Mes"].unique().tolist())
mes_label = " – ".join([MESES[meses_con_datos[0]], MESES[meses_con_datos[-1]]]) if len(meses_con_datos) > 1 else MESES[meses_con_datos[0]]
MES_ACTUAL = datetime.now(ZoneInfo("America/Bogota")).month
MESES_VENCIDOS = list(range(1, MES_ACTUAL))
meses_vencidos_label = "Sin meses vencidos" if not MESES_VENCIDOS else (
    MESES[MESES_VENCIDOS[0]] if len(MESES_VENCIDOS) == 1
    else f"{MESES[MESES_VENCIDOS[0]]} – {MESES[MESES_VENCIDOS[-1]]}"
)

ALL_P   = ["Todos"] + sorted(df_full["Proyecto"].unique().tolist())
ALL_E   = ["Todas"] + sorted(df_full["ETAPA"].unique().tolist())
ALL_M   = ["Todos"] + list(MESES.values())
ALL_MAT = ["Todos"] + sorted(df_full["MATERIAL"].unique().tolist())
ALL_EST = ["Todos"] + list(ESTADO_MAP.values())
all_ciudades = set()
if "Ciudad" in df_controles.columns:
    all_ciudades.update(df_controles["Ciudad"].dropna().tolist())
if "Ciudad" in df_full.columns:
    all_ciudades.update(df_full["Ciudad"].dropna().tolist())
ALL_CIUD = ["Todas"] + sorted(all_ciudades)

all_proyectos_controles = set()
if "Proyecto" in df_controles.columns:
    all_proyectos_controles.update(df_controles["Proyecto"].dropna().tolist())
if "Proyecto" in df_full.columns:
    all_proyectos_controles.update(df_full["Proyecto"].dropna().tolist())
ALL_PC = ["Todos"] + sorted(all_proyectos_controles)

# ── HELPERS ────────────────────────────────────────────────────────────────────
def kpi(icon, label, value, sub, css):
    return (f'<div class="kpi-card {css}"><div class="kpi-icon">{icon}</div>'
            f'<div class="kpi-label">{label}</div><div class="kpi-value">{value}</div>'
            f'<div class="kpi-sub">{sub}</div></div>')


def render_echarts(option, height=320):
    chart_id = f"echart-{uuid4().hex}"
    option_json = json.dumps(option, ensure_ascii=False)
    option_json_js = json.dumps(option_json, ensure_ascii=False)
    html = f"""
    <div id="{chart_id}" style="width:100%;height:{height}px;"></div>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
    <script>
      function reviveEchartsFunctions(value) {{
        if (Array.isArray(value)) {{
          return value.map(reviveEchartsFunctions);
        }}
        if (value && typeof value === 'object') {{
          for (const key of Object.keys(value)) {{
            value[key] = reviveEchartsFunctions(value[key]);
          }}
          return value;
        }}
        if (typeof value === 'string' && value.startsWith('__JS__') && value.endsWith('__JS__')) {{
          const fnBody = value.slice(6, -6);
          return eval('(' + fnBody + ')');
        }}
        return value;
      }}

      const chart = echarts.init(document.getElementById('{chart_id}'));
      const rawOption = {option_json_js};
      const option = reviveEchartsFunctions(JSON.parse(rawOption));
      chart.setOption(option);
      window.addEventListener('resize', () => chart.resize());
    </script>
    """
    components.html(html, height=height, scrolling=False)

def section_header(title, subtitle=""):
    subtitle_html = f'<div class="section-sub">{subtitle}</div>' if subtitle else ""
    return (
        f'<div class="section-head">'
        f'<div class="section-title">{title}</div>'
        f'{subtitle_html}'
        f'</div>'
    )


def render_page_heading(title, subtitle=""):
    subtitle_html = f'<div class="page-sub">{subtitle}</div>' if subtitle else ""
    st.markdown(
        f'<div class="page-head"><div><div class="page-title">{title}</div>{subtitle_html}</div></div>',
        unsafe_allow_html=True
    )


def dataframe_to_excel_bytes(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Consulta")
    output.seek(0)
    return output.getvalue()


def get_image_data_uri(path):
    if not path.exists():
        return ""
    ext = path.suffix.lower()
    mime_map = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }
    mime = mime_map.get(ext, "application/octet-stream")
    encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
    return f"data:{mime};base64,{encoded}"


def build_sidebar_logo_html(path):
    image_uri = get_image_data_uri(path)
    if not image_uri:
        return ""
    return (
        '<div class="sidebar-logo-wrap">'
        f'<img src="{image_uri}" alt="Cusezar" class="sidebar-logo-img" />'
        '</div>'
    )


def month_name_to_number(month_name):
    return next((k for k, v in MESES.items() if v == month_name), None)


def build_report_cutoff_text(month_num, year=None):
    report_year = year or datetime.now(ZoneInfo("America/Bogota")).year
    last_day = monthrange(report_year, month_num)[1]
    return f"{last_day} de {MESES[month_num].lower()} del {report_year}"


def rows_have_visible_values(rows):
    return any(
        any(value is not None for value in row.get("values", []))
        for row in rows
    )


def build_general_report_table_html(tab0_data, mes_num):
    materiales_map = tab0_data["material_month_maps"][mes_num]
    torre_map = tab0_data["control_month_maps"]["Torre"][mes_num]
    producto_map = tab0_data["control_month_maps"]["Producto terminado"][mes_num]
    zonas_map = tab0_data["control_month_maps"]["Zonas comunes"][mes_num]
    diseno_map = tab0_data["control_month_maps"]["Diseño"][mes_num]
    curado_map = tab0_data["control_month_maps"]["Curado"][mes_num]
    project_city_map = tab0_data["project_city_map"]
    mostrar_diseno = tab0_data["include_design"]
    acumulado_map = tab0_data["accumulated_maps"][mes_num]

    proyectos_general = sorted(
        set(materiales_map.keys()) |
        set(torre_map.keys()) |
        set(producto_map.keys()) |
        set(zonas_map.keys()) |
        set(diseno_map.keys()) |
        set(curado_map.keys()) |
        set(acumulado_map.keys())
    )

    header_cells = [
        "<th>Proyecto</th>",
        "<th>Control de Materiales</th>",
        "<th>Control de Torre</th>",
        "<th>Control de producto terminado</th>",
        "<th>Control de Zonas comunes</th>",
    ]
    if mostrar_diseno:
        header_cells.append("<th>Control de diseño</th>")
    header_cells.extend([
        "<th>Curado</th>",
        "<th>Promedio mes</th>",
        "<th>Promedio acumulado</th>",
    ])

    city_groups = {}
    for proyecto in proyectos_general:
        ciudad = project_city_map.get(normalize_project_key(proyecto), "Sin ciudad")
        city_groups.setdefault(ciudad, []).append(proyecto)

    body_rows = []
    corp_row_values = []
    for ciudad in sorted(city_groups):
        city_projects = sorted(city_groups[ciudad])
        city_row_values = []

        for proyecto in city_projects:
            row_values = [
                materiales_map.get(proyecto),
                torre_map.get(proyecto),
                producto_map.get(proyecto),
                zonas_map.get(proyecto),
            ]
            if mostrar_diseno:
                row_values.append(diseno_map.get(proyecto))
            row_values.append(curado_map.get(proyecto))

            promedio_mes = average_values(row_values)
            promedio_acumulado = acumulado_map.get(proyecto)
            city_row_values.append(row_values + [promedio_mes, promedio_acumulado])
            corp_row_values.append(row_values + [promedio_mes, promedio_acumulado])

            row_html = [f"<tr><td>{html.escape(str(proyecto))}</td>"]
            row_html.append(percent_cell_html(materiales_map.get(proyecto)))
            row_html.append(percent_cell_html(torre_map.get(proyecto)))
            row_html.append(percent_cell_html(producto_map.get(proyecto)))
            row_html.append(percent_cell_html(zonas_map.get(proyecto)))
            if mostrar_diseno:
                row_html.append(percent_cell_html(diseno_map.get(proyecto)))
            row_html.append(percent_cell_html(curado_map.get(proyecto)))
            row_html.append(percent_cell_html(promedio_mes))
            row_html.append(percent_cell_html(promedio_acumulado))
            row_html.append("</tr>")
            body_rows.append("".join(row_html))

        city_columns_avg = []
        total_columns = 8 if mostrar_diseno else 7
        for col_idx in range(total_columns):
            city_columns_avg.append(average_values(row[col_idx] for row in city_row_values))

        city_row_html = [f'<tr class="ig-summary-row"><td>{html.escape(str(ciudad))}</td>']
        for value in city_columns_avg:
            city_row_html.append(percent_cell_html(value))
        city_row_html.append("</tr>")
        body_rows.append("".join(city_row_html))

    corp_columns_avg = []
    total_columns = 8 if mostrar_diseno else 7
    for col_idx in range(total_columns):
        corp_columns_avg.append(average_values(row[col_idx] for row in corp_row_values))

    corp_row_html = ['<tr class="ig-corp-row"><td>Cusezar</td>']
    for value in corp_columns_avg:
        corp_row_html.append(percent_cell_html(value))
    corp_row_html.append("</tr>")
    body_rows.append("".join(corp_row_html))

    return (
        '<div class="ig-wrap">'
        '<div class="report-table-shell report-general-table-shell">'
        f'<table class="ig-table"><thead><tr>{"".join(header_cells)}</tr></thead>'
        f'<tbody>{"".join(body_rows)}</tbody></table></div></div>'
    )


def build_pending_ensayos_project_map(df_ensayos, month_num):
    pending_df = df_ensayos[
        (df_ensayos["Mes"] == month_num) &
        (df_ensayos["Cantidad_num"].isin([0, 0.5]))
    ].copy()
    pending_map = {}
    if pending_df.empty:
        return pending_map

    pending_df = pending_df.sort_values(
        ["Proyecto", "ETAPA", "Estado", "ENSAYO"],
        ascending=[True, True, True, True],
    )

    for proyecto, group in pending_df.groupby("Proyecto", sort=True):
        items = []
        for _, row in group.iterrows():
            etapa = str(row.get("ETAPA", "")).strip()
            ensayo = str(row.get("ENSAYO", "")).strip()
            estado = str(row.get("Estado", "")).strip()
            if not ensayo or ensayo.lower() in {"nan", "none"}:
                continue
            item = f"{etapa}: {ensayo}" if etapa and etapa.lower() not in {"nan", "none"} else ensayo
            if estado and estado.lower() not in {"nan", "none"}:
                item = f"{item} ({estado})"
            items.append(item)

        unique_items = list(dict.fromkeys(items))
        if unique_items:
            pending_map[proyecto] = "<br>".join(html.escape(item) for item in unique_items)

    return pending_map


def build_report_pending_table_html(df_ensayos, df_ctrl, month_num):
    df_ctrl_month = df_ctrl[df_ctrl["Mes"] == month_num].copy() if "Mes" in df_ctrl.columns else df_ctrl.copy()
    proyectos_ctrl = sorted(
        set(df_ctrl_month["Proyecto"].dropna().tolist()) |
        set(df_ensayos["Proyecto"].dropna().tolist())
    ) if {"Proyecto"}.issubset(df_ctrl_month.columns) and {"Proyecto"}.issubset(df_ensayos.columns) else []

    ctrl_rows = build_pending_controls_rows(df_ctrl_month, proyectos_ctrl, show_repeat_count=False)
    ctrl_map = {row["Proyecto"]: row for row in ctrl_rows}
    ens_map = build_pending_ensayos_project_map(df_ensayos, month_num)

    all_projects = sorted(set(ctrl_map.keys()) | set(ens_map.keys()))
    if not all_projects:
        return ""

    empty_html = '<span style="color:#9CA3AF;">—</span>'
    rows_html = []
    for proyecto in all_projects:
        ctrl_row = ctrl_map.get(proyecto, {})
        rows_html.append(
            "<tr>"
            f"<td>{html.escape(str(proyecto))}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{ctrl_row.get('Control de torre', empty_html)}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{ctrl_row.get('Producto terminado de torres', empty_html)}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{ctrl_row.get('Control zonas comunes', empty_html)}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{ens_map.get(proyecto, empty_html)}</td>"
            "</tr>"
        )

    return (
        '<div class="report-table-shell">'
        '<table class="rt"><thead><tr>'
        '<th>Proyecto</th>'
        '<th>Control de torre</th>'
        '<th>Producto terminado de torres</th>'
        '<th>Control zonas comunes</th>'
        '<th>Ensayos pendientes</th>'
        f'</tr></thead><tbody>{"".join(rows_html)}</tbody></table></div>'
    )


def build_report_artifacts(month_num):
    tab0_data = build_tab0_precomputed_data(EXCEL_SIGNATURE)
    report_year = datetime.now(ZoneInfo("America/Bogota")).year
    cutoff_text = build_report_cutoff_text(month_num, report_year)
    report_heatmap_height = 500
    logo_uri = get_image_data_uri(SIDEBAR_LOGO_PATH)
    logo_block = (
        '<div class="report-logo-wrap">'
        f'<img src="{logo_uri}" alt="Cusezar" class="report-logo-img" />'
        '</div>'
    ) if logo_uri else ""

    general_table_html = build_general_report_table_html(tab0_data, month_num)
    chart_specs = []
    report_sections = []

    chart_config = tab0_data["city_chart_config"]
    if chart_config:
        combo_option, combo_height = chart_config
        report_combo_option = json.loads(json.dumps(combo_option, ensure_ascii=False))
        for series in report_combo_option.get("series", []):
            if series.get("name") == "Cusezar 2026":
                series["label"] = {
                    "show": True,
                    "position": "top",
                    "formatter": "{c}%",
                    "fontFamily": "Inter, sans-serif",
                    "fontSize": 10,
                    "fontWeight": 700,
                    "color": "#FF0000",
                }
                break
        combo_id = f"report-chart-{uuid4().hex}"
        chart_specs.append({
            "id": combo_id,
            "height": combo_height,
            "option_json": json.dumps(report_combo_option, ensure_ascii=False),
        })
        report_sections.append(
            '<section class="pdf-section page-break section-chart">'
            '<div class="report-section-title">Cumplimiento plan de calidad 2026</div>'
            f'<div id="{combo_id}" class="report-chart" style="height:{combo_height}px;"></div>'
            '</section>'
        )

    ensayos_rows = build_tab2_heatmap_rows_from_summary(build_tab2_precomputed_data(EXCEL_SIGNATURE))
    if rows_have_visible_values(ensayos_rows):
        ensayos_option, _ = build_echarts_heatmap_config(ensayos_rows)
        ensayos_id = f"report-chart-{uuid4().hex}"
        chart_specs.append({
            "id": ensayos_id,
            "height": report_heatmap_height,
            "option_json": json.dumps(ensayos_option, ensure_ascii=False),
        })
        report_sections.append(
            '<section class="pdf-section page-break section-heatmap">'
            '<div class="report-section-title">Cumplimiento de ejecución de control de ensayos</div>'
            f'{heatmap_legend()}'
            f'<div id="{ensayos_id}" class="report-chart" style="height:{report_heatmap_height}px;"></div>'
            '</section>'
        )

    for area in CONTROL_AREA_OPTIONS:
        area_rows = build_heatmap_rows(df_controles, df_full, area)
        if not rows_have_visible_values(area_rows):
            continue
        area_option, _ = build_echarts_heatmap_config(area_rows)
        area_id = f"report-chart-{uuid4().hex}"
        chart_specs.append({
            "id": area_id,
            "height": report_heatmap_height,
            "option_json": json.dumps(area_option, ensure_ascii=False),
        })
        report_sections.append(
            '<section class="pdf-section page-break section-heatmap">'
            f'<div class="report-section-title">{html.escape(control_area_title(area))}</div>'
            f'{heatmap_legend()}'
            f'<div id="{area_id}" class="report-chart" style="height:{report_heatmap_height}px;"></div>'
            '</section>'
        )

    pending_table_html = build_report_pending_table_html(df_full, df_controles, month_num)
    pending_section = (
        '<section class="pdf-section page-break section-pending">'
        '<div class="report-section-title">Controles pendientes</div>'
        f'{pending_table_html}'
        '</section>'
    ) if pending_table_html else (
        '<section class="pdf-section page-break section-pending">'
        '<div class="report-section-title">Controles pendientes</div>'
        '<div class="report-empty">No se encontraron controles ni ensayos pendientes para el mes seleccionado.</div>'
        '</section>'
    )

    report_html = f"""
    <div class="pdf-report">
      <style>
        @page {{
          size: A4 landscape;
          margin: 8mm;
        }}
        .pdf-report {{
          width: 1030px;
          box-sizing: border-box;
          background: #FFFFFF;
          color: #272829;
          font-family: Inter, Arial, sans-serif;
          padding: 18px 18px 20px 18px;
          overflow: hidden;
          display: flex;
          flex-direction: column;
          align-items: center;
        }}
        .pdf-section {{
          width: 100%;
          max-width: 960px;
          box-sizing: border-box;
          margin: 0;
          padding: 0;
          display: flex;
          flex-direction: column;
          align-items: center;
          justify-content: center;
          break-inside: avoid;
          page-break-inside: avoid;
        }}
        .pdf-section.page-break {{
          page-break-before: always;
          break-before: page;
          min-height: 630px;
          padding-top: 0;
        }}
        .pdf-section.section-table {{
          min-height: 430px;
        }}
        .pdf-section.section-pending {{
          justify-content: flex-start;
          min-height: 630px;
        }}
        .report-header {{
          display: flex;
          align-items: center;
          justify-content: center;
          gap: 24px;
          min-height: 96px;
          width: 100%;
          max-width: 960px;
          margin-bottom: 8px;
          border-bottom: 1px solid #E5E9F0;
          padding-bottom: 8px;
          box-sizing: border-box;
        }}
        .report-logo-wrap {{
          display: inline-flex;
          flex: 0 0 auto;
          justify-content: center;
          align-items: center;
          padding: 12px 18px;
          background: #FF0000;
          border-radius: 18px;
        }}
        .report-logo-img {{
          display: block;
          width: 170px;
          max-width: 170px;
          height: auto;
        }}
        .report-title {{
          flex: 1 1 auto;
          font-size: 20px;
          line-height: 1.2;
          font-weight: 800;
          color: #111827;
          word-break: break-word;
        }}
        .report-section-title {{
          align-self: stretch;
          font-size: 18px;
          line-height: 1.2;
          font-weight: 800;
          color: #111827;
          margin-bottom: 10px;
        }}
        .report-section-sub {{
          font-size: 12px;
          line-height: 1.5;
          color: #6B7280;
          margin-bottom: 12px;
        }}
        .report-empty {{
          padding: 12px 14px;
          border: 1px solid #E5E9F0;
          border-radius: 12px;
          color: #6B7280;
          background: #F8FAFC;
          font-size: 12px;
        }}
        .report-table-shell {{
          width: 100%;
          max-width: 960px;
          margin: 0 auto;
          border-radius: 10px;
          border: 1px solid #E5E9F0;
          overflow: hidden;
          background: #FFFFFF;
          box-sizing: border-box;
        }}
        .report-general-table-shell {{
          max-width: 900px;
        }}
        .report-general-table-shell .ig-table {{
          font-size: 11px;
        }}
        .report-general-table-shell .ig-table th {{
          padding: 7px 5px;
          font-size: 8px;
          line-height: 1.1;
        }}
        .report-general-table-shell .ig-table td {{
          padding: 6px 5px;
          font-size: 10px;
          line-height: 1.12;
        }}
        .report-general-table-shell .ig-table td strong {{
          font-size: 10px;
        }}
        .report-chart {{
          width: 100%;
          max-width: 100%;
          overflow: hidden;
        }}
        .hml {{ display:flex; gap:8px; flex-wrap:wrap; margin-bottom:14px; align-items:center; justify-content:center; }}
        .hml span {{ font-size:11px; font-weight:600; padding:3px 10px; border-radius:20px; }}
        .rt {{ width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed; }}
        .rt th {{ background:#B5545C; padding:9px 10px; text-align:left; font-size:10px; font-weight:700; color:#FFF7F7; text-transform:uppercase; letter-spacing:.05em; border-bottom:1px solid #9F434B; white-space:normal; line-height:1.2; word-break:break-word; }}
        .rt td {{ padding:9px 10px; border-bottom:1px solid #F1D9DB; color:#6B7280; line-height:1.35; word-break:break-word; vertical-align:top; }}
        .rt td:first-child {{ background:#F8E8E8; color:#8F3942; font-weight:700; }}
        .rt tr:last-child td {{ border-bottom:none; }}
        .ig-wrap {{ display:flex; justify-content:center; }}
        .ig-table {{ width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed; }}
        .ig-table th {{ background:#B5545C; padding:8px 8px; text-align:center; font-size:10px; font-weight:700; color:#FFF7F7; text-transform:uppercase; letter-spacing:.05em; border-bottom:1px solid #9F434B; white-space:normal; line-height:1.2; word-break:break-word; }}
        .ig-table td {{ padding:7px 8px; border-bottom:1px solid #F1D9DB; color:#6B7280; text-align:center; line-height:1.15; word-break:break-word; }}
        .ig-table td:first-child {{ background:#F8E8E8; color:#8F3942; font-weight:700; text-align:center; }}
        .ig-table tr.ig-summary-row td {{ background:#F8E8E8 !important; border-top:1px solid #E8CDD0; border-bottom:1px solid #E8CDD0; font-weight:700; color:#8F3942 !important; }}
        .ig-table tr.ig-summary-row td:first-child {{ background:#F8E8E8 !important; color:#8F3942 !important; text-transform:uppercase; letter-spacing:.04em; }}
        .ig-table tr.ig-summary-row td.h100,
        .ig-table tr.ig-summary-row td.h75,
        .ig-table tr.ig-summary-row td.h50,
        .ig-table tr.ig-summary-row td.h25,
        .ig-table tr.ig-summary-row td.h0,
        .ig-table tr.ig-summary-row td.hna {{ background:#F8E8E8 !important; }}
        .ig-table tr.ig-corp-row td {{ background:#E1B7BB !important; border-top:1px solid #C89499; border-bottom:1px solid #C89499; font-weight:800; }}
        .ig-table tr.ig-corp-row td:first-child {{ background:#E1B7BB !important; color:#5E1F28 !important; text-transform:uppercase; letter-spacing:.05em; }}
        .ig-table tr.ig-corp-row td.h100,
        .ig-table tr.ig-corp-row td.h75,
        .ig-table tr.ig-corp-row td.h50,
        .ig-table tr.ig-corp-row td.h25,
        .ig-table tr.ig-corp-row td.h0,
        .ig-table tr.ig-corp-row td.hna {{ background:#E1B7BB !important; }}
        .ig-table tr:last-child td {{ border-bottom:none; }}
        .h100 {{ background:#B8E4D0; color:#2D6A4F; }}
        .h75 {{ background:#DDE8B2; color:#667A1E; }}
        .h50 {{ background:#F4E1A6; color:#A97B12; }}
        .h25 {{ background:#EEC39F; color:#A45724; }}
        .h0 {{ background:#F0C8C8; color:#8B2B2B; }}
        .hna {{ background:#F8F9FB; color:#C4CAD4; font-weight:500; font-size:11px; }}
      </style>
      <div class="report-header">
        {logo_block}
        <div class="report-title">Informe de Aplicación del plan de calidad - Corte al {html.escape(cutoff_text)}</div>
      </div>
      <section class="pdf-section section-table">
        {heatmap_legend()}
        {general_table_html}
      </section>
      {"".join(report_sections)}
      {pending_section}
    </div>
    """

    filename = f"informe_plan_calidad_{report_year}_{month_num:02d}.pdf"
    return report_html, chart_specs, filename


def render_pdf_download_button(report_html, chart_specs, filename, button_label="Informe"):
    component_id = f"pdf-report-{uuid4().hex}"
    report_html_js = json.dumps(report_html, ensure_ascii=False)
    chart_specs_js = json.dumps(chart_specs, ensure_ascii=False)
    filename_js = json.dumps(filename, ensure_ascii=False)
    button_label_js = json.dumps(button_label, ensure_ascii=False)

    component_html = f"""
    <div style="padding-top:27px;">
      <button id="{component_id}-btn" style="width:100%;background:#B5545C;color:#FFF7F7;border:1px solid #9F434B;border-radius:10px;padding:9px 12px;font-size:13px;font-weight:700;cursor:pointer;">
        {button_label}
      </button>
      <div id="{component_id}-status" style="font-family:Inter, Arial, sans-serif;font-size:11px;color:#9CA3AF;margin-top:6px;min-height:14px;"></div>
      <div id="{component_id}-host" style="position:absolute;left:-20000px;top:0;width:1030px;background:#FFFFFF;"></div>
    </div>
    <script src="https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/html2pdf.js/0.10.1/html2pdf.bundle.min.js"></script>
    <script>
      const reportHtml = {report_html_js};
      const chartSpecs = {chart_specs_js};
      const filename = {filename_js};
      const buttonLabel = {button_label_js};
      const host = document.getElementById('{component_id}-host');
      const button = document.getElementById('{component_id}-btn');
      const status = document.getElementById('{component_id}-status');

      host.innerHTML = reportHtml;

      function reviveEchartsFunctions(value) {{
        if (Array.isArray(value)) return value.map(reviveEchartsFunctions);
        if (value && typeof value === 'object') {{
          for (const key of Object.keys(value)) {{
            value[key] = reviveEchartsFunctions(value[key]);
          }}
          return value;
        }}
        if (typeof value === 'string' && value.startsWith('__JS__') && value.endsWith('__JS__')) {{
          const fnBody = value.slice(6, -6);
          return eval('(' + fnBody + ')');
        }}
        return value;
      }}

      async function waitForLibrary(checkFn, timeoutMs) {{
        const startedAt = Date.now();
        while (!checkFn()) {{
          if (Date.now() - startedAt > timeoutMs) {{
            throw new Error('Timeout cargando librerías del informe');
          }}
          await new Promise(resolve => setTimeout(resolve, 150));
        }}
      }}

      async function waitForImages(root, timeoutMs) {{
        const images = Array.from(root.querySelectorAll('img'));
        if (!images.length) return;
        await Promise.race([
          Promise.all(images.map((img) => new Promise((resolve) => {{
            if (img.complete && img.naturalWidth > 0) {{
              resolve();
              return;
            }}
            const done = () => resolve();
            img.addEventListener('load', done, {{ once: true }});
            img.addEventListener('error', done, {{ once: true }});
          }}))),
          new Promise((_, reject) => setTimeout(() => reject(new Error('Timeout cargando imágenes del informe')), timeoutMs))
        ]);
      }}

      function renderCharts() {{
        chartSpecs.forEach((spec) => {{
          const el = document.getElementById(spec.id);
          if (!el || el.dataset.rendered === '1') return;
          const chart = echarts.init(el, null, {{ renderer: 'canvas' }});
          const option = reviveEchartsFunctions(JSON.parse(spec.option_json));
          chart.setOption(option);
          el.dataset.rendered = '1';
        }});
      }}

      async function generatePdf() {{
        button.disabled = true;
        button.style.opacity = '0.75';
        status.textContent = 'Generando informe...';
        try {{
          await waitForLibrary(() => typeof window.echarts !== 'undefined', 15000);
          await waitForLibrary(() => typeof window.html2pdf !== 'undefined', 15000);
          const reportNode = host.firstElementChild;
          await waitForImages(reportNode, 15000);
          renderCharts();
          await new Promise(resolve => setTimeout(resolve, 1400));
          await html2pdf().set({{
            margin: [8, 8, 8, 8],
            filename: filename,
            image: {{ type: 'jpeg', quality: 0.98 }},
            html2canvas: {{ scale: 2, useCORS: true, backgroundColor: '#FFFFFF' }},
            jsPDF: {{ unit: 'mm', format: 'a4', orientation: 'landscape' }},
            pagebreak: {{ mode: ['css', 'legacy'] }},
          }}).from(reportNode).save();
          status.textContent = 'PDF generado.';
        }} catch (error) {{
          console.error(error);
          status.textContent = 'No fue posible generar el PDF en este momento.';
        }} finally {{
          button.disabled = false;
          button.style.opacity = '1';
          button.textContent = buttonLabel;
        }}
      }}

      button.addEventListener('click', generatePdf);
    </script>
    """
    components.html(component_html, height=76, scrolling=False)

def get_kpis(df):
    ex   = df[df["EsEjecutado"]]
    comp = int((ex["Cantidad_num"] == 1).sum())
    inc  = int((ex["Cantidad_num"] == 0.5).sum())
    no_r = int((ex["Cantidad_num"] == 0).sum())
    plan = int(len(df))
    pend = int((df["Cantidad"] == "*").sum())
    tot  = comp + inc + no_r
    tasa = round((comp + inc * 0.5) / tot * 100, 1) if tot > 0 else 0.0
    return comp, inc, no_r, plan, pend, tot, tasa

def filt(df, col, val, empty_val):
    return df if val == empty_val else df[df[col] == val]

def filt_mes(df, val):
    if val == "Todos": return df
    return df[df["Mes"].isin([k for k, v in MESES.items() if v == val])]

def hm_cls(t):
    if t >= 90: return "h100"
    if t >= 70: return "h75"
    if t >= 50: return "h50"
    if t >= 25: return "h25"
    return "h0"

def heatmap_legend():
    return f"""<div class="hml">
      <span style="background:#B8E4D0;color:#2D6A4F;">≥ 90%</span>
      <span style="background:#DDE8B2;color:#667A1E;">70–89%</span>
      <span style="background:#F4E1A6;color:#A97B12;">50–69%</span>
      <span style="background:#EEC39F;color:#A45724;">25–49%</span>
      <span style="background:#F0C8C8;color:#8B2B2B;">&lt; 25%</span>
      <span style="background:#F8F9FB;color:#C4CAD4;">Sin datos</span>
      <span style="font-size:11px;color:#9CA3AF;margin-left:4px;">· Meta: {META}%</span>
    </div>"""

HEATMAP_RANGE_COLORS = {
    "na": "#F8F9FB",
    "lt25": "#F0C8C8",
    "lt50": "#EEC39F",
    "lt70": "#F4E1A6",
    "lt90": "#DDE8B2",
    "gte90": "#B8E4D0",
}


def build_echarts_heatmap_config(rows_data, show_tooltip=True):
    month_labels = [MESES[m][:3] for m in range(1, 13)] + ["Prom."]
    col_values = [[] for _ in range(12)]
    matrix_rows = []

    for row in rows_data:
        row_vals = [v for v in row["values"] if v is not None]
        row_avg = round(sum(row_vals) / len(row_vals), 1) if row_vals else None
        display_values = []
        titles = row.get("titles", [None] * 12)
        for idx, val in enumerate(row["values"]):
            if val is not None:
                col_values[idx].append(val)
            display_values.append({"value": val, "title": titles[idx]})
        display_values.append({"value": row_avg, "title": "Promedio de la fila" if row_avg is not None else None})
        matrix_rows.append({"label": row["label"], "cells": display_values})

    avg_cells = []
    avg_row_values = []
    for idx in range(12):
        vals = col_values[idx]
        if vals:
            avg_val = round(sum(vals) / len(vals), 1)
            avg_row_values.append(avg_val)
            avg_cells.append({"value": avg_val, "title": "Promedio de la columna"})
        else:
            avg_cells.append({"value": None, "title": None})
    grand_avg = round(sum(avg_row_values) / len(avg_row_values), 1) if avg_row_values else None
    avg_cells.append({"value": grand_avg, "title": "Promedio general" if grand_avg is not None else None})
    matrix_rows.append({"label": "Promedio", "cells": avg_cells})

    data = []
    y_labels = [row["label"] for row in matrix_rows]
    for y_idx, row in enumerate(matrix_rows):
        for x_idx, cell in enumerate(row["cells"]):
            value = cell["value"]
            visual_value = -1 if value is None else round(float(value), 1)
            display_text = "—" if value is None else f"{round(float(value))}%"
            data.append([x_idx, y_idx, visual_value, display_text, cell["title"] or "Sin datos"])

    option = {
        "textStyle": {"fontFamily": "Inter, sans-serif"},
        "tooltip": {
            "show": show_tooltip,
            "position": "top",
            "formatter": """__JS__function (params) {
                const raw = params.data;
                const titulo = raw[4];
                const valor = raw[3];
                return '<b>' + params.name + '</b><br/>' + valor + '<br/>' + titulo;
            }__JS__""",
        },
        "grid": {"left": 110, "right": 18, "top": 10, "bottom": 20, "containLabel": True},
        "xAxis": {
            "type": "category",
            "position": "top",
            "data": month_labels,
            "splitArea": {"show": True},
            "axisLabel": {"color": "#6B7280", "fontFamily": "Inter, sans-serif", "fontSize": 11},
            "axisLine": {"lineStyle": {"color": "#E5E7EB"}},
        },
        "yAxis": {
            "type": "category",
            "data": y_labels,
            "inverse": True,
            "splitArea": {"show": True},
            "axisLabel": {"color": "#374151", "fontFamily": "Inter, sans-serif", "fontSize": 11},
            "axisTick": {"show": False},
            "axisLine": {"lineStyle": {"color": "#E5E7EB"}},
        },
        "visualMap": {
            "show": False,
            "min": -1,
            "max": 100,
            "dimension": 2,
            "pieces": [
                {"value": -1, "color": HEATMAP_RANGE_COLORS["na"]},
                {"min": 0, "max": 24.999, "color": HEATMAP_RANGE_COLORS["lt25"]},
                {"min": 25, "max": 49.999, "color": HEATMAP_RANGE_COLORS["lt50"]},
                {"min": 50, "max": 69.999, "color": HEATMAP_RANGE_COLORS["lt70"]},
                {"min": 70, "max": 89.999, "color": HEATMAP_RANGE_COLORS["lt90"]},
                {"min": 90, "max": 100, "color": HEATMAP_RANGE_COLORS["gte90"]},
            ],
        },
        "series": [{
            "name": "Cumplimiento",
            "type": "heatmap",
            "data": data,
            "label": {
                "show": True,
                "formatter": """__JS__function (params) {
                    const value = params.data[2];
                    const text = params.data[3];
                    if (value === -1) return '{na|' + text + '}';
                    if (value < 25) return '{lt25|' + text + '}';
                    if (value < 50) return '{lt50|' + text + '}';
                    if (value < 70) return '{lt70|' + text + '}';
                    if (value < 90) return '{lt90|' + text + '}';
                    return '{gte90|' + text + '}';
                }__JS__""",
                "rich": {
                    "na": {"color": "#C4CAD4", "fontFamily": "Inter, sans-serif", "fontSize": 10, "fontWeight": 700},
                    "lt25": {"color": "#8B2B2B", "fontFamily": "Inter, sans-serif", "fontSize": 10, "fontWeight": 700},
                    "lt50": {"color": "#A45724", "fontFamily": "Inter, sans-serif", "fontSize": 10, "fontWeight": 700},
                    "lt70": {"color": "#A97B12", "fontFamily": "Inter, sans-serif", "fontSize": 10, "fontWeight": 700},
                    "lt90": {"color": "#667A1E", "fontFamily": "Inter, sans-serif", "fontSize": 10, "fontWeight": 700},
                    "gte90": {"color": "#2D6A4F", "fontFamily": "Inter, sans-serif", "fontSize": 10, "fontWeight": 700},
                },
            },
            "itemStyle": {"borderColor": "#E5E9F0", "borderWidth": 1},
            "emphasis": {"itemStyle": {"shadowBlur": 0, "borderColor": "#CBD5E1", "borderWidth": 1}},
        }],
    }
    return option, max(280, 44 + len(y_labels) * 32)

def badge(estado):
    m = {"Completo":("bc","✅"), "Incompleto":("bi","⚠️"),
         "No Realizado":("bn","❌"), "Planeado":("bp","🔵")}
    cls, ico = m.get(estado, ("bp","🔵"))
    return f'<span class="badge {cls}">{ico} {estado}</span>'


def pending_text_color(estado):
    return {
        "No Realizado": "#8B2B2B",
        "Incompleto": "#A97B12",
        "Completo": "#2D6A4F",
        "Planeado": "#4A7BA8",
    }.get(estado, "#6B7280")

def sem_card(name, tasa, ej, crit):
    cls = "sv" if tasa >= META else "sa" if tasa >= META * 0.6 else "sr"
    dot = "#6BBF9E" if tasa >= META else "#E8C17A" if tasa >= META * 0.6 else "#D98B8B"
    return (f'<div class="sem-card {cls}"><div class="sem-name">'
            f'<span class="sem-dot" style="background:{dot}"></span>{name}</div>'
            f'<div class="sem-tasa">{tasa:.1f}%</div>'
            f'<div class="sem-detail">{ej} ejecutables · {crit} crítico{"s" if crit!=1 else ""}</div></div>')

def bar_col(t):
    return COLORS["Completo"] if t >= META else COLORS["Incompleto"] if t >= META * 0.6 else COLORS["No Realizado"]

def control_area_title(area):
    return {
        "Torre": "Control de Torres",
        "Zonas comunes": "Control de Zonas Comunes",
        "Diseño": "Control de Diseño",
        "Curado": "Control de Curado",
        "Producto terminado": "Control de Producto Terminado",
    }.get(area, f"Control de {area}")

def control_area_mask(df, area):
    if df is None or "Area" not in df.columns:
        return pd.Series([False] * len(df), index=df.index if df is not None else pd.Index([]))
    if area == "Curado":
        if "Control" not in df.columns:
            return pd.Series([False] * len(df), index=df.index)
        return (df["Area"] == "-") & (df["Control"] == "Curado")
    if area == "Producto terminado":
        return pd.Series([False] * len(df), index=df.index)
    if area == "Torre":
        if "Etapa" not in df.columns:
            return df["Area"] == area
        return (df["Area"] == area) & (~df["Etapa"].astype(str).str.contains("Producto terminado", case=False, na=False))
    return df["Area"] == area

def build_pending_controls_rows(df_ctrl, proyectos, show_repeat_count=False):
    required_cols = {"Proyecto", "Area", "Etapa", "Control", "Valor_num"}
    if df_ctrl is None or not required_cols.issubset(df_ctrl.columns):
        return []

    pending_df = df_ctrl[df_ctrl["Valor_num"].isin([0, 0.5])].copy()
    rows = []

    for proyecto in proyectos:
        proy_df = pending_df[pending_df["Proyecto"] == proyecto]

        torre_mask = (
            (proy_df["Area"] == "Torre") &
            (proy_df["Etapa"].isin(["Estructura", "Obra gris y Obra blanca"]))
        )
        producto_mask = (
            (proy_df["Area"] == "Torre") &
            (proy_df["Etapa"] == "Producto terminado")
        )
        zonas_mask = proy_df["Area"] == "Zonas comunes"

        def format_controls(subdf):
            controles = (
                subdf["Control"]
                .dropna()
                .astype(str)
                .str.strip()
            )
            controles = [c for c in controles.tolist() if c and c not in {"nan", "None"}]
            if not controles:
                return '<span style="color:#9CA3AF;">—</span>'
            if show_repeat_count:
                counts = {}
                ordered = []
                for control in controles:
                    if control not in counts:
                        counts[control] = 0
                        ordered.append(control)
                    counts[control] += 1
                formatted = [
                    f"{control} ({counts[control]})" if counts[control] >= 2 else control
                    for control in ordered
                ]
                return "<br>".join(formatted)
            controles = list(dict.fromkeys(controles))
            return "<br>".join(controles)

        rows.append({
            "Proyecto": proyecto,
            "Control de torre": format_controls(proy_df[torre_mask]),
            "Producto terminado de torres": format_controls(proy_df[producto_mask]),
            "Control zonas comunes": format_controls(proy_df[zonas_mask]),
        })

    return [
        row for row in rows
        if any(row[col] != '<span style="color:#9CA3AF;">—</span>' for col in [
            "Control de torre",
            "Producto terminado de torres",
            "Control zonas comunes",
        ])
    ]


def get_material_month_map(df_ens, month):
    required_cols = {"Proyecto", "EsEjecutado", "Mes", "Cantidad_num"}
    if df_ens is None or not required_cols.issubset(df_ens.columns):
        return {}

    result = {}
    for proyecto in sorted(df_ens["Proyecto"].dropna().unique().tolist()):
        sub = df_ens[
            (df_ens["Proyecto"] == proyecto) &
            (df_ens["EsEjecutado"]) &
            (df_ens["Mes"] == month)
        ]
        result[proyecto] = round(sub["Cantidad_num"].mean() * 100, 1) if not sub.empty else None
    return result


def month_map_from_rows(rows, month):
    return {
        row["label"]: row["values"][month - 1] if len(row["values"]) >= month else None
        for row in rows
    }


def sanitize_echarts_series(values):
    return ["-" if value is None or pd.isna(value) else round(float(value), 1) for value in values]


def build_project_city_map_from_dfs(df_ensayos, df_ctrl):
    frames = []
    for df in (df_ensayos, df_ctrl):
        if {"Proyecto", "Ciudad"}.issubset(df.columns):
            sub = df[["Proyecto", "Ciudad"]].copy()
            sub = sub.dropna(subset=["Proyecto", "Ciudad"])
            sub["Proyecto"] = sub["Proyecto"].astype(str).str.strip()
            sub["Ciudad"] = sub["Ciudad"].astype(str).str.strip()
            sub = sub[
                sub["Proyecto"].ne("") &
                sub["Ciudad"].ne("") &
                sub["Proyecto"].str.lower().ne("nan") &
                sub["Ciudad"].str.lower().ne("nan") &
                sub["Proyecto"].str.lower().ne("none") &
                sub["Ciudad"].str.lower().ne("none")
            ]
            sub["ProyectoKey"] = sub["Proyecto"].map(normalize_project_key)
            sub = sub[sub["ProyectoKey"].notna()]
            frames.append(sub)

    if not frames:
        return {}

    cities_df = pd.concat(frames, ignore_index=True).drop_duplicates()
    city_map = {}
    for proyecto_key, group in cities_df.groupby("ProyectoKey", sort=True):
        ciudades = (
            group["Ciudad"]
            .dropna()
            .astype(str)
            .str.strip()
        )
        ciudades = [
            ciudad for ciudad in ciudades.tolist()
            if ciudad and ciudad.lower() not in {"nan", "none"}
        ]
        if ciudades:
            city_map[proyecto_key] = ciudades[0]
    return city_map


def build_city_combo_chart_config_from_series(city_month_series, cusezar_2026_month_series, cusezar_2025_month_series):
    month_labels_chart = [MESES[m] for m in range(1, 13)]
    city_names = [
        ciudad for ciudad in sorted(city_month_series.keys())
        if any(value is not None and not pd.isna(value) for value in city_month_series[ciudad])
    ]
    city_palette = [
        "#D98B8B", "#E8C17A", "#7BA7D4", "#6BBF9E", "#5B8FF9", "#F6BD16",
        "#61DDAA", "#65789B", "#7262FD", "#78D3F8", "#9661BC", "#F6903D",
    ]

    has_combo_data = (
        bool(city_names) or
        any(value is not None and not pd.isna(value) for value in cusezar_2026_month_series) or
        any(value is not None and not pd.isna(value) for value in cusezar_2025_month_series)
    )
    if not has_combo_data:
        return None

    combo_option = {
        "textStyle": {"fontFamily": "Inter, sans-serif"},
        "color": city_palette + ["#B5545C"],
        "animation": False,
        "tooltip": {"trigger": "axis", "axisPointer": {"type": "cross"}},
        "legend": {
            "type": "scroll",
            "bottom": 2,
            "left": "center",
            "data": city_names + ["Cusezar 2026", "Cusezar 2025"],
            "textStyle": {"fontFamily": "Inter, sans-serif", "fontSize": 12, "color": "#6B7280"},
        },
        "grid": {"left": 45, "right": 20, "top": 20, "bottom": 72, "containLabel": True},
        "xAxis": {
            "type": "category",
            "data": month_labels_chart,
            "axisLabel": {"color": "#6B7280", "fontFamily": "Inter, sans-serif", "fontSize": 11},
            "axisLine": {"lineStyle": {"color": "#D1D5DB"}},
        },
        "yAxis": {
            "type": "value",
            "min": 0,
            "max": 100,
            "axisLabel": {"formatter": "{value}%", "color": "#6B7280", "fontFamily": "Inter, sans-serif"},
            "splitLine": {"lineStyle": {"color": "#F3F4F6"}},
        },
        "series": [
            {
                "name": ciudad,
                "type": "bar",
                "barMaxWidth": 18,
                "data": sanitize_echarts_series(city_month_series[ciudad]),
            }
            for ciudad in city_names
        ] + [{
            "name": "Cusezar 2026",
            "type": "line",
            "smooth": True,
            "connectNulls": False,
            "symbolSize": 8,
            "lineStyle": {"width": 3, "color": "#FF0000"},
            "itemStyle": {"color": "#FF0000"},
            "data": sanitize_echarts_series(cusezar_2026_month_series),
        }, {
            "name": "Cusezar 2025",
            "type": "line",
            "smooth": True,
            "connectNulls": False,
            "symbolSize": 7,
            "lineStyle": {"width": 2.5, "type": "dashed", "color": "#909AA0"},
            "itemStyle": {"color": "#909AA0"},
            "data": sanitize_echarts_series(cusezar_2025_month_series),
        }],
    }
    return combo_option, 390


def build_cusezar_2025_series(df_2025):
    if df_2025 is None or df_2025.empty:
        return [None] * 12

    company_col = next(
        (col for col in df_2025.columns if str(col).strip().casefold() in {"empresa", "compania", "compañia", "proyecto", "nombre"}),
        df_2025.columns[0] if len(df_2025.columns) else None
    )
    if company_col is None:
        return [None] * 12

    cusezar_row = df_2025[df_2025[company_col].astype(str).str.strip().str.casefold() == "cusezar"]
    if cusezar_row.empty:
        return [None] * 12

    row = cusezar_row.iloc[0]
    series = [None] * 12
    for col in df_2025.columns:
        month = normalize_month_header(col)
        if month is None:
            continue
        series[month - 1] = parse_text_value(row[col])
    return series


def build_project_accumulated_maps_from_precomputed(material_month_maps, control_month_maps, include_design):
    all_projects = set()
    for month_map in material_month_maps.values():
        all_projects.update(month_map.keys())
    for area_month_maps in control_month_maps.values():
        for month_map in area_month_maps.values():
            all_projects.update(month_map.keys())

    accumulated_by_month = {}
    for selected_month in range(1, 13):
        accumulated = {}
        for proyecto in sorted(all_projects):
            month_averages = []
            for month in range(1, selected_month + 1):
                row_values = [
                    material_month_maps[month].get(proyecto),
                    control_month_maps["Torre"][month].get(proyecto),
                    control_month_maps["Producto terminado"][month].get(proyecto),
                    control_month_maps["Zonas comunes"][month].get(proyecto),
                ]
                if include_design:
                    row_values.append(control_month_maps["Diseño"][month].get(proyecto))
                row_values.append(control_month_maps["Curado"][month].get(proyecto))

                monthly_average = average_values(row_values)
                if monthly_average is not None:
                    month_averages.append(monthly_average)

            accumulated[proyecto] = average_values(month_averages)
        accumulated_by_month[selected_month] = accumulated

    return accumulated_by_month


def build_city_month_chart_data_from_precomputed(material_month_maps, control_month_maps, include_design, project_city_map):
    all_projects = set()
    for month_map in material_month_maps.values():
        all_projects.update(month_map.keys())
    for area_month_maps in control_month_maps.values():
        for month_map in area_month_maps.values():
            all_projects.update(month_map.keys())

    city_month_values = {}
    cusezar_series = []

    for month in range(1, 13):
        month_city_values = {}
        for proyecto in sorted(all_projects):
            row_values = [
                material_month_maps[month].get(proyecto),
                control_month_maps["Torre"][month].get(proyecto),
                control_month_maps["Producto terminado"][month].get(proyecto),
                control_month_maps["Zonas comunes"][month].get(proyecto),
            ]
            if include_design:
                row_values.append(control_month_maps["Diseño"][month].get(proyecto))
            row_values.append(control_month_maps["Curado"][month].get(proyecto))

            promedio_mes = average_values(row_values)
            if promedio_mes is None:
                continue

            ciudad = project_city_map.get(normalize_project_key(proyecto), "Sin ciudad")
            month_city_values.setdefault(ciudad, []).append(promedio_mes)

        month_city_avg = {
            ciudad: average_values(values)
            for ciudad, values in month_city_values.items()
        }

        for ciudad, value in month_city_avg.items():
            city_month_values.setdefault(ciudad, [None] * 12)
            city_month_values[ciudad][month - 1] = value

        cusezar_series.append(average_values(month_city_avg.values()))

    return city_month_values, cusezar_series


@st.cache_data
def build_tab0_precomputed_data(excel_signature):
    df_ensayos, df_ctrl, df_2025 = load_data(excel_signature)
    areas = ["Torre", "Producto terminado", "Zonas comunes", "Diseño", "Curado"]

    material_month_maps = {
        month: get_material_month_map(df_ensayos, month)
        for month in range(1, 13)
    }
    control_rows = {
        area: build_heatmap_rows(df_ctrl, df_ensayos, area)
        for area in areas
    }
    control_month_maps = {
        area: {
            month: month_map_from_rows(control_rows[area], month)
            for month in range(1, 13)
        }
        for area in areas
    }
    project_city_map = build_project_city_map_from_dfs(df_ensayos, df_ctrl)
    include_design = any(
        value is not None
        for month_map in control_month_maps["Diseño"].values()
        for value in month_map.values()
    )
    accumulated_maps = build_project_accumulated_maps_from_precomputed(
        material_month_maps,
        control_month_maps,
        include_design,
    )
    city_month_series, cusezar_2026_month_series = build_city_month_chart_data_from_precomputed(
        material_month_maps,
        control_month_maps,
        include_design,
        project_city_map,
    )
    cusezar_2025_month_series = build_cusezar_2025_series(df_2025)

    return {
        "material_month_maps": material_month_maps,
        "control_month_maps": control_month_maps,
        "project_city_map": project_city_map,
        "include_design": include_design,
        "accumulated_maps": accumulated_maps,
        "city_chart_config": build_city_combo_chart_config_from_series(
            city_month_series,
            cusezar_2026_month_series,
            cusezar_2025_month_series,
        ),
    }


@st.cache_data
def build_tab2_precomputed_data(excel_signature):
    df_ensayos, _, _ = load_data(excel_signature)
    summary = df_ensayos.copy()
    summary["comp"] = (summary["Cantidad_num"] == 1).astype(int)
    summary["inc"] = (summary["Cantidad_num"] == 0.5).astype(int)
    summary["no_r"] = (summary["Cantidad_num"] == 0).astype(int)
    summary["ejecuted"] = summary["EsEjecutado"].astype(int)
    summary["planned"] = (summary["Cantidad"] == "*").astype(int)

    grouped = (
        summary.groupby(["Proyecto", "ETAPA", "MATERIAL", "Mes"], dropna=False)
        .agg(
            comp=("comp", "sum"),
            inc=("inc", "sum"),
            no_r=("no_r", "sum"),
            executed=("ejecuted", "sum"),
            planned=("planned", "sum"),
        )
        .reset_index()
    )
    return grouped


def filter_tab2_summary(summary_df, proyecto, etapa, material):
    filtered = summary_df.copy()
    if proyecto != "Todos":
        filtered = filtered[filtered["Proyecto"] == proyecto]
    if etapa != "Todas":
        filtered = filtered[filtered["ETAPA"] == etapa]
    if material != "Todos":
        filtered = filtered[filtered["MATERIAL"] == material]
    return filtered


def filter_summary_by_month(summary_df, mes):
    if mes == "Todos":
        return summary_df
    month_num = next((k for k, v in MESES.items() if v == mes), None)
    if month_num is None:
        return summary_df.iloc[0:0].copy()
    return summary_df[summary_df["Mes"] == month_num]


def build_tab2_heatmap_rows_from_summary(summary_df):
    rows_hm = []
    if summary_df.empty:
        return rows_hm

    for proyecto in sorted(summary_df["Proyecto"].dropna().unique().tolist()):
        values = []
        titles = []
        proy_df = summary_df[summary_df["Proyecto"] == proyecto]
        for month in range(1, 13):
            sub = proy_df[proy_df["Mes"] == month]
            if sub.empty:
                values.append(None)
                titles.append(None)
                continue

            comp = int(sub["comp"].sum())
            inc = int(sub["inc"].sum())
            no_r = int(sub["no_r"].sum())
            executed = int(sub["executed"].sum())
            planned = int(sub["planned"].sum())

            if executed == 0:
                values.append(None)
                titles.append("Plan." if planned > 0 else None)
                continue

            cumplimiento = round(((comp + inc * 0.5) / executed) * 100, 1)
            values.append(cumplimiento)
            titles.append(f"{comp} compl. · {inc} incompl. · {no_r} no-real · Ejecutados: {executed}")

        rows_hm.append({"label": proyecto, "values": values, "titles": titles})

    return rows_hm


def build_tab2_tasa_df_from_summary(summary_df):
    if summary_df.empty:
        return pd.DataFrame(columns=["Proyecto", "tasa", "comp", "inc", "no_r", "tot"])

    vencidos = summary_df[summary_df["Mes"].isin(MESES_VENCIDOS)].copy()
    if vencidos.empty:
        return pd.DataFrame(columns=["Proyecto", "tasa", "comp", "inc", "no_r", "tot"])

    rows = []
    for proyecto, group in vencidos.groupby("Proyecto", sort=True):
        comp = int(group["comp"].sum())
        inc = int(group["inc"].sum())
        no_r = int(group["no_r"].sum())
        tot = int((group["executed"] + group["planned"]).sum())
        tasa = round(((comp + inc * 0.5) / tot) * 100, 1) if tot > 0 else 0.0
        rows.append({
            "Proyecto": proyecto,
            "tasa": tasa,
            "comp": comp,
            "inc": inc,
            "no_r": no_r,
            "tot": tot,
        })

    return pd.DataFrame(rows).sort_values("tasa", ascending=False)


def build_tasa_df_for_selected_period(summary_df, mes):
    if mes == "Todos":
        return build_tab2_tasa_df_from_summary(summary_df)

    period_df = filter_summary_by_month(summary_df, mes)
    if period_df.empty:
        return pd.DataFrame(columns=["Proyecto", "tasa", "comp", "inc", "no_r", "tot"])

    rows = []
    for proyecto, group in period_df.groupby("Proyecto", sort=True):
        comp = int(group["comp"].sum())
        inc = int(group["inc"].sum())
        no_r = int(group["no_r"].sum())
        tot = int((group["executed"] + group["planned"]).sum())
        tasa = round(((comp + inc * 0.5) / tot) * 100, 1) if tot > 0 else 0.0
        rows.append({
            "Proyecto": proyecto,
            "tasa": tasa,
            "comp": comp,
            "inc": inc,
            "no_r": no_r,
            "tot": tot,
        })

    return pd.DataFrame(rows).sort_values("tasa", ascending=False)


@st.cache_data
def get_tab5_view_data(excel_signature, ciudad, proyecto, area, pending_month_key):
    df_ensayos, df_ctrl, _ = load_data(excel_signature)

    if ciudad != "Todas":
        if "Ciudad" in df_ctrl.columns:
            df_ctrl = df_ctrl[df_ctrl["Ciudad"] == ciudad]
        if "Ciudad" in df_ensayos.columns:
            df_ensayos = df_ensayos[df_ensayos["Ciudad"] == ciudad]

    if proyecto != "Todos":
        if "Proyecto" in df_ctrl.columns:
            df_ctrl = df_ctrl[df_ctrl["Proyecto"] == proyecto]
        if "Proyecto" in df_ensayos.columns:
            df_ensayos = df_ensayos[df_ensayos["Proyecto"] == proyecto]

    visible_until_month = MESES_VENCIDOS[-1] if MESES_VENCIDOS else 0
    rows5 = build_heatmap_rows(df_ctrl, df_ensayos, area, visible_until_month=visible_until_month)

    df_pending = df_ctrl if pending_month_key == "Todos" else df_ctrl[df_ctrl["Mes"] == pending_month_key]
    proyectos_tabla = sorted(
        set(df_pending["Proyecto"].dropna().tolist()) |
        set(df_ensayos["Proyecto"].dropna().tolist())
    ) if {"Proyecto"}.issubset(df_pending.columns) and {"Proyecto"}.issubset(df_ensayos.columns) else []

    pending_rows = build_pending_controls_rows(
        df_pending,
        proyectos_tabla,
        show_repeat_count=(pending_month_key == "Todos"),
    )

    return rows5, pending_rows


@st.cache_data
def get_tab5_available_areas(excel_signature, ciudad, proyecto):
    df_ensayos, df_ctrl, _ = load_data(excel_signature)

    if ciudad != "Todas":
        if "Ciudad" in df_ctrl.columns:
            df_ctrl = df_ctrl[df_ctrl["Ciudad"] == ciudad]
        if "Ciudad" in df_ensayos.columns:
            df_ensayos = df_ensayos[df_ensayos["Ciudad"] == ciudad]

    if proyecto != "Todos":
        if "Proyecto" in df_ctrl.columns:
            df_ctrl = df_ctrl[df_ctrl["Proyecto"] == proyecto]
        if "Proyecto" in df_ensayos.columns:
            df_ensayos = df_ensayos[df_ensayos["Proyecto"] == proyecto]

    visible_until_month = MESES_VENCIDOS[-1] if MESES_VENCIDOS else 0
    return [
        area for area in CONTROL_AREA_OPTIONS
        if build_heatmap_rows(df_ctrl, df_ensayos, area, visible_until_month=visible_until_month)
    ]


def average_values(values):
    valid_values = [float(v) for v in values if v is not None and not pd.isna(v)]
    return round(sum(valid_values) / len(valid_values), 1) if valid_values else None


def row_has_data_until_month(values, month_limit):
    if month_limit is None:
        return any(value is not None for value in values)
    if month_limit <= 0:
        return False
    return any(value is not None for value in values[:month_limit])


def percent_cell_html(value):
    if value is None or pd.isna(value):
        return '<td class="hna"><strong>—</strong></td>'
    cls = hm_cls(float(value))
    text_colors = {
        "h100": "#2D6A4F",
        "h75": "#667A1E",
        "h50": "#A97B12",
        "h25": "#A45724",
        "h0": "#8B2B2B",
    }
    color = text_colors.get(cls, "#374151")
    return f'<td class="{cls}" style="color:{color};"><strong>{float(value):.0f}%</strong></td>'

def build_heatmap_rows(df_ctrl, df_ens, area, visible_until_month=None):
    ctrl_required = {"Proyecto", "Area", "Etapa", "Control", "Mes", "Valor_num"}
    ens_required = {"Proyecto", "ETAPA", "MATERIAL", "ENSAYO", "Mes", "Cantidad_num"}

    if df_ctrl is None:
        df_ctrl = pd.DataFrame(columns=list(ctrl_required) + ["Valor"])
    if df_ens is None:
        df_ens = pd.DataFrame(columns=list(ens_required))

    for col in ctrl_required | {"Valor"}:
        if col not in df_ctrl.columns:
            df_ctrl[col] = pd.Series(dtype="object")
    for col in ens_required:
        if col not in df_ens.columns:
            df_ens[col] = pd.Series(dtype="object")

    rows = []
    ctrl_filtered = df_ctrl[control_area_mask(df_ctrl, area)].copy()
    if area == "Producto terminado":
        proyectos_ens = set()
        for etapa, material, ensayo in PRODUCTO_TERMINADO_ENSAYOS:
            mask = (
                (df_ens["ETAPA"] == etapa) &
                (df_ens["MATERIAL"] == material) &
                (df_ens["ENSAYO"] == ensayo)
            )
            proyectos_ens.update(df_ens.loc[mask, "Proyecto"].dropna().tolist())

        proyectos_ctrl = set()
        for area_v, etapa_v, control_v in PRODUCTO_TERMINADO_CONTROLES:
            mask = (
                (df_ctrl["Area"] == area_v) &
                (df_ctrl["Etapa"] == etapa_v) &
                (df_ctrl["Control"] == control_v)
            )
            proyectos_ctrl.update(df_ctrl.loc[mask, "Proyecto"].dropna().tolist())

        proyectos = sorted(proyectos_ens | proyectos_ctrl)
    else:
        proyectos = sorted(
            set(ctrl_filtered["Proyecto"].dropna().tolist()) |
            set(df_ens["Proyecto"].dropna().tolist())
        )

    for proyecto in proyectos:
        ctrl_proy = ctrl_filtered[ctrl_filtered["Proyecto"] == proyecto]
        ens_parts = []
        ctrl_parts = []
        if area == "Producto terminado":
            for etapa, material, ensayo in PRODUCTO_TERMINADO_ENSAYOS:
                mask = (
                    (df_ens["Proyecto"] == proyecto) &
                    (df_ens["ETAPA"] == etapa) &
                    (df_ens["MATERIAL"] == material) &
                    (df_ens["ENSAYO"] == ensayo)
                )
                ens_parts.append(df_ens.loc[mask, ["Mes", "Cantidad_num"]].rename(columns={"Cantidad_num": "Valor_num"}))

            for area_v, etapa_v, control_v in PRODUCTO_TERMINADO_CONTROLES:
                mask = (
                    (df_ctrl["Proyecto"] == proyecto) &
                    (df_ctrl["Area"] == area_v) &
                    (df_ctrl["Etapa"] == etapa_v) &
                    (df_ctrl["Control"] == control_v)
                )
                ctrl_parts.append(df_ctrl.loc[mask, ["Mes", "Valor_num"]])

        extras = pd.concat(ens_parts + ctrl_parts, ignore_index=True) if (ens_parts or ctrl_parts) else pd.DataFrame(columns=["Mes", "Valor_num"])
        values = []
        titles = []

        for m in range(1, 13):
            vals_ctrl = ctrl_proy.loc[ctrl_proy["Mes"] == m, "Valor_num"].dropna()
            vals_extra = extras.loc[extras["Mes"] == m, "Valor_num"].dropna()
            vals = pd.concat([vals_ctrl, vals_extra], ignore_index=True)
            if area == "Curado":
                raw_vals = ctrl_proy.loc[ctrl_proy["Mes"] == m, "Valor"].dropna()
                parsed_vals = [parse_text_value(v) for v in raw_vals]
                parsed_vals = [v for v in parsed_vals if v is not None]
                if not parsed_vals:
                    values.append(None)
                    titles.append(None)
                    continue
                t = round(parsed_vals[-1], 1)
                values.append(t)
                titles.append("Valor registrado")
            elif vals.empty:
                values.append(None)
                titles.append(None)
            else:
                t = round(vals.mean() * 100, 1)
                values.append(t)
                titles.append(f"Promedio de {len(vals)} valores")
        if row_has_data_until_month(values, visible_until_month):
            rows.append({"label": proyecto, "values": values, "titles": titles})

    return rows

NAV_OPTIONS = {
    "Informe General": "⌂  Informe General",
    "Ensayos": "◫  Ensayos",
    "Controles": "☷  Controles",
    "Consulta de Ensayos": "⌕  Consulta de Ensayos",
}

with st.sidebar:
    logo_html = build_sidebar_logo_html(SIDEBAR_LOGO_PATH)
    if logo_html:
        st.markdown(logo_html, unsafe_allow_html=True)
    st.markdown('<div class="sidebar-brand">Calidad</div>', unsafe_allow_html=True)
    st.markdown('<div class="sidebar-sub">Panel de navegación</div>', unsafe_allow_html=True)
    current_page = st.radio(
        "Navegación",
        list(NAV_OPTIONS.keys()),
        format_func=lambda key: NAV_OPTIONS[key],
        label_visibility="collapsed",
    )
    st.markdown(
        f'<div class="sidebar-footer"><strong>Ultima actualización</strong>{EXCEL_LAST_UPDATE_TEXT}</div>',
        unsafe_allow_html=True
    )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 0
# ══════════════════════════════════════════════════════════════════════════════
if current_page == "Informe General":
    render_page_heading("Informe General", "Resumen consolidado del tablero de calidad")
    default_mes_general = MESES[MESES_VENCIDOS[-1]] if MESES_VENCIDOS else list(MESES.values())[0]
    default_mes_general_idx = ALL_M.index(default_mes_general) if default_mes_general in ALL_M else 1

    gcol, bcol, _ = st.columns([1.2, 0.85, 3.95])
    with gcol:
        sel0_mes = st.selectbox(
            "Mes",
            ALL_M[1:],
            index=max(0, default_mes_general_idx - 1),
            key="t0m",
        )

    mes0_num = month_name_to_number(sel0_mes)
    report_html, report_chart_specs, report_filename = build_report_artifacts(mes0_num)
    with bcol:
        render_pdf_download_button(report_html, report_chart_specs, report_filename, button_label="Informe")

    st.markdown(section_header(f"Informe General de {sel0_mes}", f"Resumen consolidado por proyecto"), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)
    st.markdown(heatmap_legend(), unsafe_allow_html=True)
    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    tab0_data = build_tab0_precomputed_data(EXCEL_SIGNATURE)
    materiales_map = tab0_data["material_month_maps"][mes0_num]
    torre_map = tab0_data["control_month_maps"]["Torre"][mes0_num]
    producto_map = tab0_data["control_month_maps"]["Producto terminado"][mes0_num]
    zonas_map = tab0_data["control_month_maps"]["Zonas comunes"][mes0_num]
    diseno_map = tab0_data["control_month_maps"]["Diseño"][mes0_num]
    curado_map = tab0_data["control_month_maps"]["Curado"][mes0_num]
    project_city_map = tab0_data["project_city_map"]

    mostrar_diseno = tab0_data["include_design"]
    acumulado_map = tab0_data["accumulated_maps"][mes0_num]
    proyectos_general = sorted(
        set(materiales_map.keys()) |
        set(torre_map.keys()) |
        set(producto_map.keys()) |
        set(zonas_map.keys()) |
        set(diseno_map.keys()) |
        set(curado_map.keys()) |
        set(acumulado_map.keys())
    )

    header_cells = [
        "<th>Proyecto</th>",
        "<th>Control de Materiales</th>",
        "<th>Control de Torre</th>",
        "<th>Control de producto terminado</th>",
        "<th>Control de Zonas comunes</th>",
    ]
    if mostrar_diseno:
        header_cells.append("<th>Control de diseño</th>")
    header_cells.extend([
        "<th>Curado</th>",
        "<th>Promedio mes</th>",
        "<th>Promedio acumulado</th>",
    ])

    city_groups = {}
    for proyecto in proyectos_general:
        ciudad = project_city_map.get(normalize_project_key(proyecto), "Sin ciudad")
        city_groups.setdefault(ciudad, []).append(proyecto)

    body_rows = []
    corp_row_values = []
    for ciudad in sorted(city_groups):
        city_projects = sorted(city_groups[ciudad])
        city_row_values = []

        for proyecto in city_projects:
            row_values = [
                materiales_map.get(proyecto),
                torre_map.get(proyecto),
                producto_map.get(proyecto),
                zonas_map.get(proyecto),
            ]
            if mostrar_diseno:
                row_values.append(diseno_map.get(proyecto))
            row_values.append(curado_map.get(proyecto))

            promedio_mes = average_values(row_values)
            promedio_acumulado = acumulado_map.get(proyecto)
            city_row_values.append(row_values + [promedio_mes, promedio_acumulado])
            corp_row_values.append(row_values + [promedio_mes, promedio_acumulado])

            row_html = [f"<tr><td>{proyecto}</td>"]
            row_html.append(percent_cell_html(materiales_map.get(proyecto)))
            row_html.append(percent_cell_html(torre_map.get(proyecto)))
            row_html.append(percent_cell_html(producto_map.get(proyecto)))
            row_html.append(percent_cell_html(zonas_map.get(proyecto)))
            if mostrar_diseno:
                row_html.append(percent_cell_html(diseno_map.get(proyecto)))
            row_html.append(percent_cell_html(curado_map.get(proyecto)))
            row_html.append(percent_cell_html(promedio_mes))
            row_html.append(percent_cell_html(promedio_acumulado))
            row_html.append("</tr>")
            body_rows.append("".join(row_html))

        city_columns_avg = []
        total_columns = 8 if mostrar_diseno else 7
        for col_idx in range(total_columns):
            city_columns_avg.append(average_values(
                row[col_idx] for row in city_row_values
            ))

        city_row_html = [f'<tr class="ig-summary-row"><td>{ciudad}</td>']
        for value in city_columns_avg:
            city_row_html.append(percent_cell_html(value))
        city_row_html.append("</tr>")
        body_rows.append("".join(city_row_html))

    corp_columns_avg = []
    total_columns = 8 if mostrar_diseno else 7
    for col_idx in range(total_columns):
        corp_columns_avg.append(average_values(
            row[col_idx] for row in corp_row_values
        ))

    corp_row_html = ['<tr class="ig-corp-row"><td>Cusezar</td>']
    for value in corp_columns_avg:
        corp_row_html.append(percent_cell_html(value))
    corp_row_html.append("</tr>")
    body_rows.append("".join(corp_row_html))

    st.markdown(
        f'<div class="ig-wrap">'
        f'<div style="width:100%;overflow-x:auto;border-radius:10px;border:1px solid #E5E9F0;max-width:1180px;">'
        f'<table class="ig-table"><thead><tr>{"".join(header_cells)}</tr></thead>'
        f'<tbody>{"".join(body_rows)}</tbody></table></div></div>',
        unsafe_allow_html=True
    )
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
    st.markdown(section_header("Cumplimiento plan de calidad 2026", "Barras por ciudad y línea de Cusezar con el promedio mensual de todas las ciudades con dato"), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)

    chart_config = tab0_data["city_chart_config"]
    if chart_config:
        combo_option, combo_height = chart_config
        render_echarts(combo_option, height=combo_height)
    else:
        st.info("ℹ️ No hay datos suficientes para construir el gráfico mensual por ciudad.")
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1
# ══════════════════════════════════════════════════════════════════════════════
if current_page == "Ensayos":
    render_page_heading("Ensayos", "Seguimiento general de ensayos por proyecto, etapa y mes")
    c1, c2, c3 = st.columns(3)
    sel_proy  = c1.selectbox("Proyecto", ALL_P,  key="t1p")
    sel_etapa = c2.selectbox("Etapa",    ALL_E,  key="t1e")
    sel_mes   = c3.selectbox("Mes",      ALL_M,  key="t1m")

    df1 = filt(filt(filt_mes(df_full, sel_mes), "ETAPA", sel_etapa, "Todas"), "Proyecto", sel_proy, "Todos")
    tab1_summary = filter_tab2_summary(
        build_tab2_precomputed_data(EXCEL_SIGNATURE),
        sel_proy,
        sel_etapa,
        "Todos",
    )
    tab1_summary = filter_summary_by_month(tab1_summary, sel_mes)

    comp, inc, no_r, plan, pend, tot, tasa = get_kpis(df1)
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.markdown(kpi("📋","Planeados",         f"{plan:,}", f"Pendientes: {pend:,}",                              "kp-blue"),   unsafe_allow_html=True)
    k2.markdown(kpi("✅","Completos",         f"{comp:,}", f"{comp/tot*100:.1f}% del ejecutable" if tot else "—","kp-green"),  unsafe_allow_html=True)
    k3.markdown(kpi("⚠️","Incompletos (Sin subir a ACC)",       f"{inc:,}",  f"{inc/tot*100:.1f}% del ejecutable"  if tot else "—","kp-yellow"), unsafe_allow_html=True)
    k4.markdown(kpi("❌","No Realizados",     f"{no_r:,}", f"{no_r/tot*100:.1f}% del ejecutable" if tot else "—","kp-red"),    unsafe_allow_html=True)
    k5.markdown(kpi("📈","Tasa Cumplimiento", f"{tasa}%",  f"Meta: ≥ {META}%",                                  "kp-slate"),  unsafe_allow_html=True)
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    # ── Donut ──
    d1, d2 = st.columns([1, 1.5])
    with d1:
        st.markdown(section_header("Distribución por Estado", "Total de registros en el plan"), unsafe_allow_html=True)
        st.markdown('<div class="dash-card">', unsafe_allow_html=True)
        ec = df1["Estado"].value_counts().reset_index()
        ec.columns = ["Estado","n"]
        donut_states = [estado for estado in ["Planeado", "Completo", "Incompleto", "No Realizado"] if estado in ec["Estado"].tolist()]
        ec_plot = (ec.set_index("Estado")
                     .reindex(donut_states)
                     .reset_index())
        donut_option = {
            "textStyle": {"fontFamily": "Inter, sans-serif"},
            "tooltip": {"trigger": "item", "formatter": "{b}<br/>{c} ({d}%)"},
            "legend": {
                "type": "scroll",
                "bottom": 2,
                "left": "center",
                "icon": "circle",
                "textStyle": {"fontFamily": "Inter, sans-serif", "fontSize": 12, "color": "#6B7280"},
            },
            "series": [{
                "type": "pie",
                "radius": ["48%", "72%"],
                "center": ["50%", "40%"],
                "avoidLabelOverlap": True,
                "label": {"show": True, "formatter": "{d}%", "fontFamily": "Inter, sans-serif", "fontSize": 11, "color": "#6B7280"},
                "labelLine": {"show": True, "length": 10, "length2": 8},
                "itemStyle": {"borderColor": "#FFFFFF", "borderWidth": 2},
                "data": [
                    {"name": row["Estado"], "value": int(row["n"]), "itemStyle": {"color": COLORS[row["Estado"]]}}
                    for _, row in ec_plot.iterrows()
                ],
            }],
            "graphic": [{
                "type": "text",
                "left": "center",
                "top": "40%",
                "style": {
                    "text": f"{len(df1):,}",
                    "fontSize": 18,
                    "fontWeight": 700,
                    "fontFamily": "Inter, sans-serif",
                    "fill": "#111827",
                    "align": "center",
                    "verticalAlign": "middle",
                },
            }],
        }
        render_echarts(donut_option, height=340)
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Tasa por proyecto ──
    with d2:
        tasa_subtitle = (
            f"% de completos sobre el total planeado en meses vencidos ({meses_vencidos_label}) · Meta: {META}%"
            if sel_mes == "Todos"
            else f"% de completos sobre el total planeado en {sel_mes} · Meta: {META}%"
        )
        st.markdown(section_header("Tasa de Cumplimiento por Proyecto", tasa_subtitle), unsafe_allow_html=True)
        st.markdown('<div class="dash-card">', unsafe_allow_html=True)
        if not tab1_summary.empty:
            t_df = build_tasa_df_for_selected_period(tab1_summary, sel_mes)
            tasa_option = {
                "textStyle": {"fontFamily": "Inter, sans-serif"},
                "tooltip": {
                    "trigger": "item",
                    "formatter": """__JS__function (params) {
                        const d = params.data;
                        return '<b>' + d.proyecto + '</b><br/>Cumplimiento: ' + d.value.toFixed(1) + '%' +
                               '<br/>Completos: ' + d.comp +
                               '<br/>Incompletos: ' + d.inc +
                               '<br/>No realizados: ' + d.no_r +
                               '<br/>Total plan: ' + d.tot;
                    }__JS__""",
                },
                "grid": {"left": 120, "right": 35, "top": 24, "bottom": 20, "containLabel": True},
                "xAxis": {
                    "type": "value",
                    "min": 0,
                    "max": 115,
                    "axisLabel": {"formatter": "{value}%", "color": "#6B7280", "fontFamily": "Inter, sans-serif"},
                    "splitLine": {"show": True, "lineStyle": {"color": "#F3F4F6"}},
                },
                "yAxis": {
                    "type": "category",
                    "data": t_df["Proyecto"].tolist(),
                    "axisLabel": {"color": "#374151", "fontFamily": "Inter, sans-serif"},
                    "axisTick": {"show": False},
                },
                "series": [{
                    "type": "bar",
                    "data": [
                        {
                            "value": float(row.tasa),
                            "proyecto": row.Proyecto,
                            "comp": int(row.comp),
                            "inc": int(row.inc),
                            "no_r": int(row.no_r),
                            "tot": int(row.tot),
                            "itemStyle": {"color": bar_col(row.tasa)},
                        }
                        for _, row in t_df.iterrows()
                    ],
                    "barWidth": 18,
                    "label": {
                        "show": True,
                        "position": "right",
                        "formatter": "{c}%",
                        "fontFamily": "Inter, sans-serif",
                        "fontSize": 11,
                        "color": "#6B7280",
                    },
                    "markLine": {
                        "silent": True,
                        "symbol": "none",
                        "lineStyle": {"type": "dashed", "color": COLORS["Planeado"], "width": 1.5},
                        "label": {
                            "show": True,
                            "formatter": f"Meta {META}%",
                            "color": COLORS["Planeado"],
                            "fontFamily": "Inter, sans-serif",
                            "fontSize": 10,
                        },
                        "data": [{"xAxis": META}],
                    },
                }],
            }
            render_echarts(tasa_option, height=max(240, 55 + len(t_df) * 22))
        else:
            st.info("ℹ️ No hay datos suficientes para calcular la tasa de cumplimiento con los filtros aplicados.")
        st.markdown('</div>', unsafe_allow_html=True)

    # ── Heatmap ──
    st.markdown(section_header('Cumplimiento de ejecución de control de ensayos'), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)
    st.markdown(heatmap_legend(), unsafe_allow_html=True)
    rows_hm = build_tab2_heatmap_rows_from_summary(tab1_summary)
    heatmap_option, heatmap_height = build_echarts_heatmap_config(rows_hm)
    render_echarts(heatmap_option, height=heatmap_height)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Línea temporal ──
    st.markdown(section_header("Ensayos por Mes — 2026", "Líneas sólidas = ejecutados por estado · Punteada = total planeado del mes"), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)
    mp = (df1.groupby("Mes").size()
            .reindex(range(1,13), fill_value=0).reset_index())
    mp.columns = ["Mes","n"]
    line_option = {
        "textStyle": {"fontFamily": "Inter, sans-serif"},
        "color": [COLORS["Planeado"], COLORS["Completo"], COLORS["Incompleto"], COLORS["No Realizado"]],
        "tooltip": {"trigger": "axis"},
        "legend": {
            "type": "scroll",
            "bottom": 2,
            "left": "center",
            "textStyle": {"fontFamily": "Inter, sans-serif", "fontSize": 12, "color": "#6B7280"},
        },
        "grid": {"left": 35, "right": 15, "top": 20, "bottom": 68, "containLabel": True},
        "xAxis": {
            "type": "category",
            "data": [MESES[m] for m in range(1, 13)],
            "axisLabel": {"color": "#6B7280", "fontFamily": "Inter, sans-serif", "fontSize": 11},
            "axisLine": {"lineStyle": {"color": "#D1D5DB"}},
        },
        "yAxis": {
            "type": "value",
            "axisLabel": {"color": "#6B7280", "fontFamily": "Inter, sans-serif"},
            "splitLine": {"lineStyle": {"color": "#F3F4F6"}},
        },
        "series": [
            {
                "name": "Plan del mes",
                "type": "line",
                "smooth": True,
                "symbolSize": 7,
                "lineStyle": {"width": 2, "type": "dashed"},
                "data": mp["n"].astype(int).tolist(),
            },
            {
                "name": "Completo",
                "type": "line",
                "smooth": True,
                "symbolSize": 8,
                "areaStyle": {"color": "rgba(107,191,158,0.15)"},
                "data": (df1[df1["EsEjecutado"] & (df1["Estado"]=="Completo")]
                           .groupby("Mes").size()
                           .reindex(range(1,13), fill_value=0)
                           .astype(int).tolist()),
            },
            {
                "name": "Incompleto",
                "type": "line",
                "smooth": True,
                "symbolSize": 8,
                "data": (df1[df1["EsEjecutado"] & (df1["Estado"]=="Incompleto")]
                           .groupby("Mes").size()
                           .reindex(range(1,13), fill_value=0)
                           .astype(int).tolist()),
            },
            {
                "name": "No Realizado",
                "type": "line",
                "smooth": True,
                "symbolSize": 8,
                "data": (df1[df1["EsEjecutado"] & (df1["Estado"]=="No Realizado")]
                           .groupby("Mes").size()
                           .reindex(range(1,13), fill_value=0)
                           .astype(int).tolist()),
            },
        ],
    }
    render_echarts(line_option, height=340)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Ensayos pendientes ──
    st.markdown(section_header("Ensayos pendientes", "Ensayos con valor 0 o 0,5 según los filtros aplicados"), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)
    pending_ensayos_df = df1[
        df1["Cantidad_num"].isin([0, 0.5])
    ][["Proyecto", "ETAPA", "MesNombre", "ENSAYO", "Estado"]].copy()
    pending_ensayos_df.columns = ["Proyecto", "Etapa", "Mes", "Ensayo pendiente", "Estado"]

    if not pending_ensayos_df.empty:
        pending_ensayos_df = pending_ensayos_df.sort_values(
            ["Proyecto", "Etapa", "Mes", "Estado", "Ensayo pendiente"],
            ascending=[True, True, True, True, True],
        )
        rows_pending = "".join(
            f"<tr><td>{r.Proyecto}</td><td>{r.Etapa}</td><td>{r.Mes}</td><td>{r['Ensayo pendiente']}</td><td>{badge(r.Estado)}</td></tr>"
            for _, r in pending_ensayos_df.iterrows()
        )
        st.markdown(
            f'<div style="overflow-x:auto;border-radius:10px;border:1px solid #E5E9F0;max-height:420px;overflow-y:auto;">'
            f'<table class="rt"><thead><tr>'
            f'<th>Proyecto</th><th>Etapa</th><th>Mes</th><th>Ensayo pendiente</th><th>Estado</th>'
            f'</tr></thead><tbody>{rows_pending}</tbody></table></div>',
            unsafe_allow_html=True
        )
    else:
        st.markdown('<div class="ok-note">✅ No hay ensayos pendientes con los filtros aplicados.</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4
# ══════════════════════════════════════════════════════════════════════════════
if current_page == "Consulta de Ensayos":
    render_page_heading("Consulta de Ensayos", "Filtra y encuentra exactamente qué ensayos aplican según proyecto, mes y estado")
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    q1, q2, q3 = st.columns(3)
    sel4_proy  = q1.selectbox("Proyecto",  ALL_P,   key="t4p")
    sel4_etapa = q1.selectbox("Etapa",     ALL_E,   key="t4e")
    sel4_mes   = q2.selectbox("Mes",       ALL_M,   key="t4m")
    sel4_est   = q3.selectbox("Estado",    ALL_EST, key="t4est")
    buscar     = q3.text_input("🔎 Buscar por nombre de ensayo",
                               placeholder="Ej: resistencia, fraguado, granulometría...")

    df4 = df_full.copy()
    if sel4_proy  != "Todos": df4 = df4[df4["Proyecto"] == sel4_proy]
    if sel4_etapa != "Todas": df4 = df4[df4["ETAPA"]    == sel4_etapa]
    if sel4_mes   != "Todos": df4 = df4[df4["Mes"].isin([k for k,v in MESES.items() if v == sel4_mes])]
    if sel4_est   != "Todos": df4 = df4[df4["Estado"]   == sel4_est]
    if buscar:                df4 = df4[df4["ENSAYO"].str.contains(buscar, case=False, na=False)]

    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)
    st.markdown(section_header("Resultados de la Consulta"), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)

    if not df4.empty:
        disp = df4[["Proyecto","ETAPA","MATERIAL","ENSAYO","NTC","FRECUENCIA","MesNombre","Estado"]].copy()
        disp.columns = ["Proyecto","Etapa","Material","Ensayo","NTC","Frecuencia","Mes","Estado"]
        col_dl4, _ = st.columns([1, 5])
        with col_dl4:
            st.download_button("⬇ Descargar Excel",
                               dataframe_to_excel_bytes(disp),
                               "consulta_ensayos.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="dl4")
        prev = disp.head(50)
        rows4 = "".join(
            f"<tr><td>{r.Proyecto}</td><td>{r.Etapa}</td><td>{r.Material}</td>"
            f"<td>{r.Ensayo}</td><td>{r.NTC}</td>"
            f"<td style='max-width:180px;white-space:normal;font-size:11px;color:#9CA3AF'>{r.Frecuencia}</td>"
            f"<td>{r.Mes}</td><td>{badge(r.Estado)}</td></tr>"
            for _, r in prev.iterrows())
        st.markdown(f'<div class="card-sub">Mostrando {min(50,len(disp))} de {len(disp):,} registros.</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div style="overflow-x:auto;border-radius:10px;border:1px solid #E5E9F0;max-height:480px;overflow-y:auto;">'
            f'<table class="rt"><thead><tr><th>Proyecto</th><th>Etapa</th><th>Material</th>'
            f'<th>Ensayo</th><th>NTC</th><th>Frecuencia</th><th>Mes</th><th>Estado</th></tr></thead>'
            f'<tbody>{rows4}</tbody></table></div>',
            unsafe_allow_html=True)
    else:
        st.info("ℹ️ No se encontraron ensayos con los filtros aplicados.")
    st.markdown('</div>', unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5
# ══════════════════════════════════════════════════════════════════════════════
if current_page == "Controles":
    render_page_heading("Controles", "Consulta el avance mensual de controles por ciudad, proyecto y tipo de control")
    st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)

    c5a, c5b, c5c = st.columns(3)
    sel5_ciud = c5a.selectbox("Ciudad", ALL_CIUD, key="t5c")
    sel5_proy = c5b.selectbox("Proyecto", ALL_PC, key="t5p")
    available_control_areas = get_tab5_available_areas(EXCEL_SIGNATURE, sel5_ciud, sel5_proy)
    area_options = available_control_areas if available_control_areas else ["Sin datos"]
    current_area_value = st.session_state.get("t5a")
    default_area = current_area_value if current_area_value in area_options else area_options[0]
    sel5_area = c5c.selectbox(
        "Control",
        area_options,
        index=area_options.index(default_area),
        key="t5a",
        disabled=not available_control_areas,
    )
    sel5_pend_mes = "Todos"
    pending_month_key = "Todos"

    tab5_error = None
    rows5 = []
    pending_rows = []
    if available_control_areas:
        try:
            rows5, pending_rows = get_tab5_view_data(
                EXCEL_SIGNATURE,
                sel5_ciud,
                sel5_proy,
                sel5_area,
                pending_month_key,
            )
        except Exception as exc:
            tab5_error = str(exc)

    title5 = "Controles" if sel5_area == "Sin datos" else control_area_title(sel5_area)
    st.markdown(section_header(title5, "Promedio mensual con 12 meses fijos. Los registros sin valor no se incluyen en el cálculo."), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)
    st.markdown(heatmap_legend(), unsafe_allow_html=True)

    if tab5_error:
        st.warning("No fue posible construir el heatmap de controles con la estructura actual del Excel.")
    elif rows5:
        heatmap5_option, heatmap5_height = build_echarts_heatmap_config(rows5, show_tooltip=False)
        render_echarts(heatmap5_option, height=heatmap5_height)
    elif not available_control_areas:
        st.info("ℹ️ No hay controles con información disponible para los filtros y meses vencidos.")
    else:
        st.info("ℹ️ No se encontraron controles con los filtros aplicados.")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown(section_header("Controles pendientes", "Número entre parentesis (n), cantidad de controles acumulados"), unsafe_allow_html=True)
    st.markdown('<div class="dash-card">', unsafe_allow_html=True)

    pend_col, _ = st.columns([1.2, 4.8])
    with pend_col:
        sel5_pend_mes = st.selectbox("Mes", ALL_M, key="t5_pending_mes")
    pending_month_key = "Todos" if sel5_pend_mes == "Todos" else next((k for k, v in MESES.items() if v == sel5_pend_mes), "Todos")
    if available_control_areas:
        try:
            _, pending_rows = get_tab5_view_data(
                EXCEL_SIGNATURE,
                sel5_ciud,
                sel5_proy,
                sel5_area,
                pending_month_key,
            )
            tab5_error = None
        except Exception as exc:
            tab5_error = str(exc)
            pending_rows = []
    else:
        pending_rows = []

    if tab5_error:
        st.warning("No fue posible construir la tabla de pendientes de controles con la estructura actual del Excel.")
    elif pending_rows:
        rows_html = "".join(
            f"<tr>"
            f"<td>{row['Proyecto']}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{row['Control de torre']}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{row['Producto terminado de torres']}</td>"
            f"<td style='white-space:normal;line-height:1.45;'>{row['Control zonas comunes']}</td>"
            f"</tr>"
            for row in pending_rows
        )
        st.markdown(
            f'<div style="overflow-x:auto;border-radius:10px;border:1px solid #E5E9F0;max-height:420px;overflow-y:auto;">'
            f'<table class="rt"><thead><tr>'
            f'<th>Proyecto</th>'
            f'<th>Control de torre</th>'
            f'<th>Producto terminado de torres</th>'
            f'<th>Control zonas comunes</th>'
            f'</tr></thead><tbody>{rows_html}</tbody></table></div>',
            unsafe_allow_html=True
        )
    else:
        st.info("ℹ️ No se encontraron controles pendientes con los filtros aplicados.")
    st.markdown('</div>', unsafe_allow_html=True)
